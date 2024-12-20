#!/usr/bin/env python3

#===============================================================================
#
# gen_notifs.py - Generate billing notifications for each PI for month/year.
#
# ARGS:
#   1st: the BillingConfig spreadsheet.
#
# SWITCHES:
#   --billing_details_file: Location of the BillingDetails.xlsx file (default=look in BillingRoot/<year>/<month>)
#   --accounting_file: Location of accounting file (overrides BillingConfig.xlsx)
#   --billing_root:    Location of BillingRoot directory (overrides BillingConfig.xlsx)
#                      [default if no BillingRoot in BillingConfig.xlsx or switch given: CWD]
#   --year:            Year of snapshot requested. [Default is this year]
#   --month:           Month of snapshot requested. [Default is last month]
#   --pi_sheets:       Put Billing sheets from PI-specific BillingNotifications workbooks in
#                        the BillingAggregate workbook (default=False).
#
# OUTPUT:
#   BillingNotification spreadsheets for each PI
#     in BillingRoot/<year>/<month>/GBSCBilling-<pi_tag>.<year>-<month>.xlsx
#   Various messages about current processing status to STDOUT.
#
# ASSUMPTIONS:
#
# AUTHOR:
#   Keith Bettinger
#
#==============================================================================

#=====
#
# IMPORTS
#
#=====
import argparse
from collections import defaultdict
import time
import os
import re
import sys

import openpyxl
import openpyxl.styles
import openpyxl.utils
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
import json  # For 'pickling' dicts

# Simulate an "include billing_common.py".
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
exec(compile(open(os.path.join(SCRIPT_DIR, "billing_common.py"), "rb").read(), os.path.join(SCRIPT_DIR, "billing_common.py"), 'exec'))

#=====
#
# CONSTANTS
#
#=====
# From billing_common.py
global BILLING_DETAILS_PREFIX
global BILLING_AGGREGATE_PREFIX
global BILLING_NOTIFS_SHEET_COLUMNS
global BILLING_AGGREG_SHEET_COLUMNS
global BILLING_NOTIFS_PREFIX
global CONSULTING_HOURS_FREE
global CONSULTING_TRAVEL_RATE_DISCOUNT
global ACCOUNT_PREFIXES
global SUBDIR_RAWDATA
global SUBDIR_INVOICES
global BASE_STORAGE_SIZE
global EXCEL_MAX_ROWS

#=====
#
# GLOBALS
#
#=====

#
# For make_format(), a data structure to save all the dictionaries and resulting Format objects
#  which were created for a given workbook.
#
# Data Structure: dict with workbooks as keys, and values of [(property_dict, Format)*]
FORMAT_PROPS_PER_WORKBOOK = defaultdict(list)

#
# These globals are data structures read in from BillingConfig workbook.
#

# List of pi_tags.
pi_tag_list = []

# Mapping from usernames to list of [date, pi_tag].
username_to_pi_tag_dates = defaultdict(list)

# Mapping from usernames to a list of [email, full_name].
username_to_user_details = defaultdict(list)

# Mapping from pi_tags to list of [first_name, last_name, email].
pi_tag_to_names_email = defaultdict(list)

# Mapping from accounts to list of [pi_tag, %age].
account_to_pi_tag_pctages = defaultdict(list)

# Mapping from folders to list of [pi_tag, %age].
folder_to_pi_tag_pctages = defaultdict(list)

#
# These globals are data structures used to write the BillingNotification workbooks.
#

# Mapping from pi_tag to cluster account status.
pi_tag_to_cluster_acct_status = dict()

# Mapping from pi_tag to cloud account status.
pi_tag_to_cloud_acct_status = dict()

# Mapping from pi_tag to consulting account status.
pi_tag_to_consulting_acct_status = dict()

# Mapping from pi_tag to string for their affiliate status ('Stanford', 'Affiliate', 'External').
pi_tag_to_affiliation = dict()

# Mapping from pi_tag to list of [folder, size, %age].
pi_tag_to_folder_sizes = defaultdict(list)

# Mapping from pi_tag to list of [account, list of [username, cpu_core_hrs, %age]].
pi_tag_to_account_username_cpus = defaultdict(list)

# Mapping from pi_tag to list of [date, username, job_name, account, cpu_core_hrs, jobID, %age].
pi_tag_to_job_details = defaultdict(list)

# Mapping from pi_tag to list of [username, date_added, date_removed, %age].
pi_tag_to_user_details = defaultdict(list)

# Mapping from pi_tag to list of [storage_charge, computing_charge, cloud_charge, consulting_charge, total_charge].
pi_tag_to_charges = defaultdict(list)

# Mapping from pi_tag to list of [iLab service request ID, iLab service request name, iLab service request owner].
pi_tag_to_iLab_info = defaultdict(list)

# Mapping from pi_tag to set of (cloud account, %age) tuples.
pi_tag_to_cloud_account_pctages = defaultdict(set)

# Mapping from cloud account to set of cloud project IDs (several per project possible in this set).
cloud_account_to_cloud_projects = defaultdict(set)

# Mapping from cloud account to cloud account name
cloud_account_to_account_names = dict()

# Mapping from (cloud project ID, cloud account) to lists of (platform, account, description, dates, quantity, UOM, charge) tuples.
cloud_project_account_to_cloud_details = defaultdict(list)

# Mapping from (cloud project ID, cloud account) to total charge.
cloud_project_account_to_total_charges = defaultdict(float)

# Mapping from cloud project number to cloud project ID (1-to-1 mapping).
cloud_projnum_to_cloud_project = dict()

## Bioinformatics Consulting:

# Mapping from pi_tag to list of [date, summary, notes, consultant, hours, billable_hours].
pi_tag_to_consulting_details = defaultdict(list)

# Mapping from pi_tag to list of (date, summary, consultant, hours, billable_hours].
pi_tag_to_consulting_charges = defaultdict(list)

#=====
#
# FUNCTIONS
#
#=====
# From billing_common.py
global from_timestamp_to_excel_date
global from_excel_date_to_timestamp
global from_timestamp_to_date_string
global from_excel_date_to_date_string
global from_ymd_date_to_timestamp
global from_timestamp_to_datetime
global from_datetime_to_timestamp
global from_datetime_to_date_string
global sheet_get_named_column
global filter_by_dates
global argparse_get_parent_parser
global argparse_get_year_month
global argparse_get_billingroot_billingconfig
global get_subdirectory
global rowcol_to_a1_cell

#
# This function takes an arbitrary number of dicts with formatting properties in them as defined in xlswriter,
# creates a NamedStyle and applies the formatting to it, adds the format to the given workbook, and returns it.
#
# This function caches the ones it creates per workbook, so if a format is requested more than once,
#  it will simply return the previously created Format and not make a new one.
#
def make_format(wkbk, *prop_dicts):

    # Define the final property dict.
    final_prop_dict = dict()
    # Combine all the input dicts into the final dict.
    list(map(lambda d: final_prop_dict.update(d), prop_dicts))

    # Get the list of (prop_dict, Format)s for this workbook.
    prop_dict_format_list = FORMAT_PROPS_PER_WORKBOOK.setdefault(wkbk, [])

    for (prop_dict, wkbk_format) in prop_dict_format_list:
        # Is final_prop_dict already in the list?
        if final_prop_dict == prop_dict:
            # Yes: return the associated Format object.
            format_obj = wkbk_format
            break
    else:
        # Nope: new prop_dict, therefore we must make a new Format object.
        format_obj = openpyxl.styles.NamedStyle(json.dumps(final_prop_dict))
        wkbk.add_named_style(format_obj)

        # Create objects for the format
        font      = openpyxl.styles.Font()
        border    = openpyxl.styles.Border()
        alignment = openpyxl.styles.Alignment()
        number_format = ''

        # Go through the keys in the formatting dict and generate the right objects as part of the NamedStyle.
        for key in final_prop_dict.keys():

            # Font-related formats
            if   key == 'font_size':  font.size = final_prop_dict[key]
            elif key == 'font_name':  font.name = final_prop_dict[key]
            elif key == 'font_color': font.color = final_prop_dict[key]
            elif key == 'bold':       font.bold = final_prop_dict[key]
            elif key == 'italic':     font.italic = final_prop_dict[key]
            elif key == 'vertAlign':  font.vertAlign = final_prop_dict[key]
            elif key == 'underline':
                value = final_prop_dict[key]
                if isinstance(value, bool):
                    if value:
                        font.underline = 'single'
                    else:
                        font.underline = None
                else:
                    font.underline = value
            elif key == 'font_strikeout': font.strike = final_prop_dict[key]

            # Border-related formats (UNSUPPORTED:, diagonal=, diagonal_direction=, outline=, vertical=, horizontal=
            elif key == 'left':   border.left   = openpyxl.styles.Side(border_style=final_prop_dict[key])
            elif key == 'right':  border.right  = openpyxl.styles.Side(border_style=final_prop_dict[key])
            elif key == 'top':    border.top    = openpyxl.styles.Side(border_style=final_prop_dict[key])
            elif key == 'bottom': border.bottom = openpyxl.styles.Side(border_style=final_prop_dict[key])
            elif key == 'left_color':
                if border.left is not None:
                    border.left.color = final_prop_dict[key]
                else:
                    border.left = openpyxl.styles.Side(color=final_prop_dict[key])
            elif key == 'right_color':
                if border.right is not None:
                    border.right.color = final_prop_dict[key]
                else:
                    border.right = openpyxl.styles.Side(color=final_prop_dict[key])
            elif key == 'top_color':
                if border.top is not None:
                    border.top.color = final_prop_dict[key]
                else:
                    border.top = openpyxl.styles.Side(color=final_prop_dict[key])
            elif key == 'bottom_color':
                if border.bottom is not None:
                    border.bottom.color = final_prop_dict[key]
                else:
                    border.bottom = openpyxl.styles.Side(color=final_prop_dict[key])

            # Alignment-related formats
            elif key == 'align':     alignment.horizontal = final_prop_dict[key]
            elif key == 'valign':
                value = final_prop_dict[key]
                if   value == 'vcenter':      alignment.vertical = 'center'
                elif value == 'vjustify':     alignment.vertical = 'justify'
                elif value == 'vdistributed': alignment.vertical = 'distributed'
                else:                         alignment.vertical = value
            elif key == 'text_wrap': alignment.wrap_text = final_prop_dict[key]
            elif key == 'shrink':    alignment.shrink_to_fit = final_prop_dict[key]
            elif key == 'indent':    alignment.indent = final_prop_dict[key]

            # Number format
            elif key == 'num_format': number_format = final_prop_dict[key]

            # Oops, how did I get here!
            else:
                print("make_format(): Ignoring unknown key " + key, file=sys.stderr)

        # Assign the formatting subobjects to the main formatting object
        format_obj.font = font
        format_obj.border = border
        format_obj.alignment = alignment
        format_obj.number_format = number_format

        # Save the prop_dict and Format object for later use.
        prop_dict_format_list.append((final_prop_dict, format_obj))

    return format_obj


# This function creates some formats in a BillingNotification workbook,
# creates the necessary sheets, and writes the column headers in the sheets.
# It also makes the Billing sheet the active sheet when it is opened in Excel.
def init_billing_notifs_wkbk(wkbk):

    global BOLD_FORMAT
    global DATE_FORMAT
    global FLOAT_FORMAT
    global INT_FORMAT
    global MONEY_FORMAT
    global BOLD_MONEY_FORMAT
    global PERCENT_FORMAT

    # Create formats for use within the workbook.
    BOLD_FORMAT    = make_format(wkbk, {'bold' : True})
    DATE_FORMAT    = make_format(wkbk, {'num_format' : 'mm/dd/yy'})
    INT_FORMAT     = make_format(wkbk, {'num_format' : '0'})
    FLOAT_FORMAT   = make_format(wkbk, {'num_format' : '0.0'})
    MONEY_FORMAT   = make_format(wkbk, {'num_format' : '$#,##0.00'})
    BOLD_MONEY_FORMAT = make_format(wkbk, {'num_format' : '$#,##0.00', 'bold' : True})
    PERCENT_FORMAT = make_format(wkbk, {'num_format' : '0%'})

    # Control the size of the Workbook when it opens
    view = [openpyxl.workbook.views.BookView(windowWidth=18140, windowHeight=30000)]
    wkbk.views = view

    sheet_name_to_sheet = dict()

    for sheet_name in BILLING_NOTIFS_SHEET_COLUMNS:

        #sheet = wkbk.add_worksheet(sheet_name)
        sheet = wkbk.create_sheet(sheet_name)
        for col in range(0, len(BILLING_NOTIFS_SHEET_COLUMNS[sheet_name])):
            sheet.cell(1, col+1, BILLING_NOTIFS_SHEET_COLUMNS[sheet_name][col]).style = BOLD_FORMAT

        sheet_name_to_sheet[sheet_name] = sheet

    # Remove "Sheet"
    wkbk.remove(wkbk["Sheet"])

    # Make the Billing sheet the active one.
    wkbk.active = sheet_name_to_sheet['Billing']

    return sheet_name_to_sheet


# This function creates a bold format in a BillingAggregate workbook,
# creates the necessary sheets, and writes the column headers in the sheets.
# It also makes the Totals sheet the active sheet when it is opened in Excel.
def init_billing_aggreg_wkbk(wkbk, pi_tag_list):

    # Control the size of the Workbook when it opens
    view = [openpyxl.workbook.views.BookView(windowWidth=18140, windowHeight=30000)]
    wkbk.views = view

    bold_format = make_format(wkbk, {'bold' : True})

    # Remove "Sheet"
    wkbk.remove(wkbk["Sheet"])

    sheet_name_to_sheet = dict()

    for sheet_name in BILLING_AGGREG_SHEET_COLUMNS:

        sheet = wkbk.create_sheet(sheet_name)
        for col in range(0, len(BILLING_AGGREG_SHEET_COLUMNS[sheet_name])):
            sheet.cell(1, col+1, BILLING_AGGREG_SHEET_COLUMNS[sheet_name][col]).style = bold_format

        sheet_name_to_sheet[sheet_name] = sheet

    if args.pi_sheets:
        # Make a sheet for each PI.
        for pi_tag in sorted(pi_tag_list):

            sheet = wkbk.create_sheet(pi_tag)
            sheet_name_to_sheet[pi_tag] = sheet

    # Make the Aggregate sheet the active one.
    wkbk.active = sheet_name_to_sheet['Totals']

    return sheet_name_to_sheet


# This function scans the username_to_pi_tag_dates dict to create a list of [pi_tag, %age]s
# for the PIs that the given user was working for on the given date.
def get_pi_tags_for_username_by_date(username, date_timestamp):

    # Add PI Tag to the list if the given date is after date_added, but before date_removed.

    pi_tag_list = []

    pi_tag_dates = username_to_pi_tag_dates.get(username)
    if pi_tag_dates is not None:

        date_excel = from_timestamp_to_datetime(date_timestamp)

        for (pi_tag, date_added, date_removed, pctage) in pi_tag_dates:
            if date_added <= date_excel and (date_removed == '' or date_removed is None or date_excel < date_removed):
                pi_tag_list.append([pi_tag, pctage])

    return pi_tag_list


# Creates all the data structures used to write the BillingNotification workbook.
# The overall goal is to mimic the tables of the notification sheets so that
# to build the table, all that is needed is to print out one of these data structures.
def build_global_data(wkbk, begin_month_timestamp, end_month_timestamp):

    pis_sheet      = wkbk["PIs"]
    folders_sheet  = wkbk["Folders"]
    users_sheet    = wkbk["Users"]
    accounts_sheet = wkbk["Accounts"]

    begin_month_datetime = from_timestamp_to_datetime(begin_month_timestamp)
    end_month_datetime   = from_timestamp_to_datetime(end_month_timestamp)

    #
    # Create list of pi_tags.
    #
    global pi_tag_list

    pi_tag_list = list(sheet_get_named_column(pis_sheet, "PI Tag"))
    # Remove all empty cells from the end of the pi_tag_list
    while pi_tag_list[-1] is None:
        pi_tag_list = pi_tag_list[:-1]

    #
    # Create mapping from pi_tag to a list of PI name and email.
    #
    global pi_tag_to_names_email

    pi_first_names = sheet_get_named_column(pis_sheet, "PI First Name")
    pi_last_names  = sheet_get_named_column(pis_sheet, "PI Last Name")
    pi_emails      = sheet_get_named_column(pis_sheet, "PI Email")

    pi_details_list = list(zip(pi_first_names, pi_last_names, pi_emails))

    pi_tag_to_names_email = dict(list(zip(pi_tag_list, pi_details_list)))

    #
    # Organize data from the Cloud sheet, if present.
    #
    cloud_sheet = wkbk["Cloud Accounts"]

    #
    # Create mapping from pi_tag to (cloud account, %age) tuples from the BillingConfig PIs sheet.
    # Create mapping from cloud account to account names
    #
    global pi_tag_to_cloud_account_pctages
    global cloud_account_to_account_names

    cloud_platforms   = sheet_get_named_column(cloud_sheet, "Platform")
    cloud_pi_tags     = sheet_get_named_column(cloud_sheet, "PI Tag")
    cloud_accounts    = sheet_get_named_column(cloud_sheet, "Billing Account Number")
    cloud_account_names = sheet_get_named_column(cloud_sheet, "Billing Account Name")
    cloud_pctages     = sheet_get_named_column(cloud_sheet, "%age")

    cloud_dates_added = sheet_get_named_column(cloud_sheet, "Date Added")
    cloud_dates_remvd = sheet_get_named_column(cloud_sheet, "Date Removed")

    cloud_rows = filter_by_dates(list(zip(cloud_platforms, cloud_pi_tags,
                                          cloud_accounts, cloud_account_names, cloud_pctages)),
                                 list(zip(cloud_dates_added, cloud_dates_remvd)),
                                 begin_month_datetime, end_month_datetime)

    #for (pi_tag, project, projnum, projid, account, pctage) in cloud_rows:
    for (platform, pi_tag, account, acct_name, pctage) in cloud_rows:

        # Only Google Cloud is supported by automated billing (for now)
        if platform != "Google": continue

        # Associate the project name and percentage to be charged with the pi_tag.
        pi_tag_to_cloud_account_pctages[pi_tag].add((account, pctage))

        # Associate the account name with the account
        cloud_account_to_account_names[account] = acct_name

    #
    # Create mapping from pi_tags to a string denoting affiliation (Stanford/Affiliate/External).
    #
    global pi_tag_to_affiliation

    affiliation_column = sheet_get_named_column(pis_sheet, "Affiliation")

    pi_tag_to_affiliation = dict(list(zip(pi_tag_list, affiliation_column)))

    #
    # Filter pi_tag_list for PIs active in the current month.
    #
    pi_dates_added   = sheet_get_named_column(pis_sheet, "Date Added")
    pi_dates_removed = sheet_get_named_column(pis_sheet, "Date Removed")

    pi_tags_and_dates_added = list(zip(pi_tag_list, pi_dates_added, pi_dates_removed))

    for (pi_tag, date_added, date_removed) in pi_tags_and_dates_added:

        # Exit loop if pi_tag is None:
        if pi_tag is None:
            break

        # Convert the datetimes to timestamps.
        date_added_timestamp = from_datetime_to_timestamp(date_added)
        if date_removed != '' and date_removed is not None:
            date_removed_timestamp = from_datetime_to_timestamp(date_removed)
        else:
            date_removed_timestamp = end_month_timestamp + 1  # Not in this month.

        # If the date added is AFTER the end of this month, or
        #  the date removed is BEFORE the beginning of this month,
        # then remove the pi_tag from the list.
        if date_added_timestamp >= end_month_timestamp:

            print(" *** Ignoring PI %s: added after this month on %s" % (pi_tag_to_names_email[pi_tag][1],
                                                                         from_datetime_to_date_string(date_added)), file=sys.stderr)
            pi_tag_list.remove(pi_tag)

        elif date_removed_timestamp < begin_month_timestamp:

            print(" *** Ignoring PI %s: removed before this month on %s" % (pi_tag_to_names_email[pi_tag][1],
                                                                            from_datetime_to_date_string(date_removed)), file=sys.stderr)
            pi_tag_list.remove(pi_tag)

    #
    # Create mapping from usernames to a list of user details.
    #
    global username_to_user_details

    usernames  = sheet_get_named_column(users_sheet, "Username")
    emails     = sheet_get_named_column(users_sheet, "Email")
    full_names = sheet_get_named_column(users_sheet, "Full Name")

    username_details_rows = list(zip(usernames, emails, full_names))

    for (username, email, full_name) in username_details_rows:
        username_to_user_details[username] = [email, full_name]

    #
    # Create mapping from usernames to a list of pi_tag/dates.
    #
    global username_to_pi_tag_dates

    pi_tags       = sheet_get_named_column(users_sheet, "PI Tag")
    dates_added   = sheet_get_named_column(users_sheet, "Date Added")
    dates_removed = sheet_get_named_column(users_sheet, "Date Removed")
    pctages       = sheet_get_named_column(users_sheet, "%age")

    username_rows = list(zip(usernames, pi_tags, dates_added, dates_removed, pctages))

    for (username, pi_tag, date_added, date_removed, pctage) in username_rows:
        username_to_pi_tag_dates[username].append([pi_tag, date_added, date_removed, pctage])

    #
    # Create mapping from pi_tags to a list of [username, date_added, date_removed]
    #
    global pi_tag_to_user_details

    for username in username_to_pi_tag_dates:

        pi_tag_date_list = username_to_pi_tag_dates[username]

        for (pi_tag, date_added, date_removed, pctage) in pi_tag_date_list:
            pi_tag_to_user_details[pi_tag].append([username, date_added, date_removed, pctage])

    global pi_tag_to_iLab_info

    pi_tags         = sheet_get_named_column(pis_sheet, "PI Tag")
    serv_req_ids    = sheet_get_named_column(pis_sheet, "iLab Service Request ID")
    serv_req_names  = sheet_get_named_column(pis_sheet, "iLab Service Request Name")
    serv_req_owners = sheet_get_named_column(pis_sheet, "iLab Service Request Owner")

    iLab_info_rows = list(zip(pi_tags, serv_req_ids, serv_req_names, serv_req_owners))

    for (pi_tag, serv_req_id, serv_req_name, serv_req_owner) in iLab_info_rows:
        pi_tag_to_iLab_info[pi_tag] = [serv_req_id, serv_req_name, serv_req_owner]

    #
    # Create mapping from account to list of pi_tags and %ages.
    #
    global account_to_pi_tag_pctages

    accounts = sheet_get_named_column(accounts_sheet, "Account")
    pi_tags  = sheet_get_named_column(accounts_sheet, "PI Tag")
    pctages  = sheet_get_named_column(accounts_sheet, "%age")

    dates_added   = sheet_get_named_column(accounts_sheet, "Date Added")
    dates_removed = sheet_get_named_column(accounts_sheet, "Date Removed")

    account_rows = filter_by_dates(list(zip(accounts, pi_tags, pctages)), list(zip(dates_added, dates_removed)),
                                   begin_month_datetime, end_month_datetime)

    for (account, pi_tag, pctage) in account_rows:
        account_to_pi_tag_pctages[account].append([pi_tag, pctage])

    # Add pi_tags prefixed by ACCOUNT_PREFIXES to list of accounts for PI.
    for pi_tag in pi_tag_list:
        account_to_pi_tag_pctages[pi_tag].append([pi_tag, 1.0])

        for prefix in ACCOUNT_PREFIXES:
            account_to_pi_tag_pctages["%s_%s" % (prefix,pi_tag)].append([pi_tag, 1.0])

    #
    # Create mapping from folder to list of pi_tags and %ages.
    #
    global folder_to_pi_tag_pctages

    # Get the Folders from PI Sheet
    folders = sheet_get_named_column(pis_sheet, "PI Folder")
    pi_tags = sheet_get_named_column(pis_sheet, "PI Tag")
    pctages = [1.0] * len(folders)

    dates_added   = sheet_get_named_column(pis_sheet, "Date Added")
    dates_removed = sheet_get_named_column(pis_sheet, "Date Removed")

    # Add the Folders from Folder sheet
    folders += sheet_get_named_column(folders_sheet, "Folder")
    pi_tags += sheet_get_named_column(folders_sheet, "PI Tag")
    pctages += sheet_get_named_column(folders_sheet, "%age")

    dates_added   += sheet_get_named_column(folders_sheet, "Date Added")
    dates_removed += sheet_get_named_column(folders_sheet, "Date Removed")

    folder_rows = filter_by_dates(list(zip(folders, pi_tags, pctages)), list(zip(dates_added, dates_removed)),
                                  begin_month_datetime, end_month_datetime)

    for (folder, pi_tag, pctage) in folder_rows:

        # Account for multiple folders separated by commas.
        pi_folder_list = folder.split(',')

        for pi_folder in pi_folder_list:
            folder_to_pi_tag_pctages[pi_folder].append([pi_tag, pctage])

    #
    # Create mappings from pi_tags to statuses for cluster, cloud, and consulting.
    #
    pi_tags = sheet_get_named_column(pis_sheet, "PI Tag")

    global pi_tag_to_cluster_acct_status
    cluster_statuses = sheet_get_named_column(pis_sheet, "Cluster?")

    pi_tag_to_cluster_acct_status = dict(list(zip(pi_tags, cluster_statuses)))


    global pi_tag_to_cloud_acct_status
    cloud_statuses = sheet_get_named_column(pis_sheet, "Google Cloud?")

    pi_tag_to_cloud_acct_status = dict(list(zip(pi_tags, cloud_statuses)))


    global pi_tag_to_consulting_acct_status
    consulting_statuses = sheet_get_named_column(pis_sheet, "BaaS?")

    pi_tag_to_consulting_acct_status = dict(list(zip(pi_tags, consulting_statuses)))


# Reads the particular rate requested from the Rates sheet of the BillingConfig workbook.
def get_rate(wkbk, rate_type):

    #rates_sheet = wkbk.sheet_by_name('Rates')
    rates_sheet = wkbk["Rates"]

    types   = sheet_get_named_column(rates_sheet, 'Type')
    amounts = sheet_get_named_column(rates_sheet, 'Amount')

    for (type, amount) in zip(types, amounts):
        if type == rate_type:
            return amount
    else:
        return None


def get_rate_a1_cell(wkbk, rate_type):

    rates_sheet = wkbk["Rates"]

    header_row = rates_sheet.iter_cols(min_row=1, max_row=1, values_only=True)

    # Find the column numbers for 'Type' and 'Amount'.
    type_col = -1
    amt_col = -1
    idx = 1
    for col_name in header_row:
        if col_name[0] == 'Type':
            type_col = idx
            if amt_col != -1: break  # Leave loop if we have both
        elif col_name[0] == 'Amount':
            amt_col = idx
            if type_col != -1: break  # Leave loop if we have both
        idx += 1

    if type_col == -1 or amt_col == -1:
        print("get_rate_a1_cell: Can't find Type/Amount headers (%d, %d)" % (type_col, amt_col), file=sys.stderr)
        return None

    # Get column of 'Types'.
    types = rates_sheet.iter_rows(min_row=2, min_col=type_col, max_col=type_col, values_only=True)

    # When you find the row with rate_string, return the Amount col and this row.
    idx = 2
    for row in types:
        for col in row:
            if col == rate_type:
                return 'Rates!%s' % rowcol_to_a1_cell(idx + 1, amt_col, True, True)
        idx += 1
    else:
        return 0.0


def get_rate_amount_and_a1_cell_from_prefix(billing_config_wkbk, service_str, tier_str, subservice_str, affiliation_str):

    if service_str == "Local HPC Storage" or service_str == "Local Computing":

        tier_string = "%s Tier" % (tier_str.capitalize())

        # Start building rate string with service string and tier string
        rate_string = "%s (%s)" % (service_str, tier_string)

        # If there is a subservice string, add that to rate string
        if subservice_str is not None and subservice_str != "":
            rate_string += ": %s" % subservice_str

    # Finish rate string with the affiliation string
    rate_string += " - %s" % affiliation_str.capitalize()

    rate_amount  = get_rate(billing_config_wkbk, rate_string)
    rate_a1_cell = get_rate_a1_cell(billing_config_wkbk, rate_string)

    return rate_amount, rate_a1_cell


# Reads the Storage sheet of the BillingDetails workbook given, and populates
# the pi_tag_to_folder_sizes dict with the folder measurements for each PI.
def read_storage_sheet(wkbk):

    global pi_tag_to_folder_sizes

    storage_sheet = wkbk["Storage"]

    for (date, timestamp, folder, size, used, inodes_quota, inodes_used) in storage_sheet.iter_rows(min_row=2, values_only=True):
        # List of [pi_tag, %age] pairs.
        pi_tag_pctages = folder_to_pi_tag_pctages[folder]

        for (pi_tag, pctage) in pi_tag_pctages:
            pi_tag_to_folder_sizes[pi_tag].append([folder, size, pctage])


# Reads the Computing sheet of the BillingDetails workbook given, and populates
# the account_to_pi_tag_cpus, pi_tag_to_account_username_cpus, and pi_tag_to_job_details dicts.
def read_computing_sheet(wkbk):

    global pi_tag_to_job_details

    #computing_sheet = wkbk.sheet_by_name("Computing")
    computing_sheet = wkbk["Computing"]

    if args.cpu_time_unit == 'cpu-hours':
        cpu_time_denom = 3600.0
    elif args.cpu_time_unit == 'cpu-days':
        cpu_time_denom = 86400.0
    else:
        print("Arg 'cpu_time_unit' has unknown value {args.cpu_time_unit", file=sys.stderr)

    sheet_number = 1

    while True:

        for (job_date, job_timestamp, job_username, job_name, account, node, cores, wallclock, jobID) in \
            computing_sheet.iter_rows(min_row=2, values_only=True):

            # Calculate CPU time units for job.
            cpu_core_time = cores * wallclock / cpu_time_denom   # wallclock is in seconds.

            # Rename this variable for easier understanding.
            account = account.lower()

            if account != '':
                job_pi_tag_pctage_list = account_to_pi_tag_pctages[account]
            else:
                # No account means credit the job to the user's lab.
                job_pi_tag_pctage_list = get_pi_tags_for_username_by_date(job_username, job_timestamp)

            if len(job_pi_tag_pctage_list) == 0:
                print("   *** No PI associated with job ID %d, user %s, account %s" % (jobID, job_username, account))
                continue

            # Distribute this job's CPU-hrs amongst pi_tags by %ages.
            for (pi_tag, pctage) in job_pi_tag_pctage_list:

                # This list is [account, list of [username, cpu_core_hrs, %age]].
                account_username_cpu_list = pi_tag_to_account_username_cpus.get(pi_tag)

                # If pi_tag has an existing list of account/username/CPUs:
                if account_username_cpu_list is not None:

                    # Find if account for job is in list of account/CPUs for this pi_tag.
                    for pi_username_cpu_pctage_list in account_username_cpu_list:

                        (pi_account, pi_username_cpu_pctage_list) = pi_username_cpu_pctage_list

                        # If the account we are looking at is the one from the present job:
                        if pi_account == account:

                            # Find job username in list for account:
                            for username_cpu in pi_username_cpu_pctage_list:
                                if job_username == username_cpu[0]:
                                    username_cpu[1] += cpu_core_time
                                    break
                            else:
                                pi_username_cpu_pctage_list.append([job_username, cpu_core_time, pctage])

                            # Leave account_username_cpu_list loop.
                            break

                    else:
                        # No matching account in pi_tag list -- add a new one to the list.
                        account_username_cpu_list.append([account, [[job_username, cpu_core_time, pctage]]])

                # Else start a new account/CPUs list for the pi_tag.
                else:
                    pi_tag_to_account_username_cpus[pi_tag] = [[account, [[job_username, cpu_core_time, pctage]]]]

                #
                # Save job details for pi_tag.
                #
                new_job_details = [job_date, job_username, job_name, account, node, cpu_core_time, jobID, pctage]
                pi_tag_to_job_details[pi_tag].append(new_job_details)

        sheet_number += 1

        try:
            computing_sheet = wkbk["Computing %d" % sheet_number]
        except:
            break  # No more computing sheets: exit the while True loop.


# Read the Cloud sheet from the BillingDetails workbook.
def read_cloud_sheet(wkbk):

    cloud_sheet = wkbk["Cloud"]

    for (platform, account, project, description, dates, quantity, uom, charge) in cloud_sheet.iter_rows(min_row=2, values_only=True):

        # If project is of the form "<project name>(<project-id>)" or "<project name>[<project-id>]", get the "<project-id>".
        if project is not None:
            project_re = re.search("[(\[]([a-z0-9-:.]+)[\])]", project)
            if project_re is not None:
                project = project_re.group(1)
            else:
                pass  # If no parens, use the original project name.


        # Save the project that the account line item is for.
        cloud_account_to_cloud_projects[account].add(project)

        # Save the cloud item in a list of charges for that PI.
        cloud_project_account_to_cloud_details[(project, account)].append((platform, description, dates, quantity, uom, charge))

        # Accumulate the total cost of a project.
        cloud_project_account_to_total_charges[(project, account)] += float(charge)


# Reads the Consulting sheet of the BillingDetails workbook.
def read_consulting_sheet(wkbk):

    #consulting_sheet = wkbk.sheet_by_name('Consulting')
    consulting_sheet = wkbk["Consulting"]

    dates       = sheet_get_named_column(consulting_sheet, 'Date')
    pi_tags     = sheet_get_named_column(consulting_sheet, 'PI Tag')
    hours       = sheet_get_named_column(consulting_sheet, 'Hours')
    travel_hours= sheet_get_named_column(consulting_sheet, 'Travel Hours')
    consultants = sheet_get_named_column(consulting_sheet, 'Participants')
    clients     = sheet_get_named_column(consulting_sheet, 'Clients')
    summaries   = sheet_get_named_column(consulting_sheet, 'Summary')
    notes       = sheet_get_named_column(consulting_sheet, 'Notes')
    cumul_hours = sheet_get_named_column(consulting_sheet, 'Cumul Hours')

    consulting_details = list(zip(dates, pi_tags, hours, travel_hours, consultants, clients, summaries, notes, cumul_hours))

    for (date, pi_tag, hours, travel_hours, consultant, client, summary, notes, cumul_hours) in consulting_details:

        if travel_hours is None:  travel_hours = 0

        # Save the consulting item in a list of details for each PI.
        pi_tag_to_consulting_details[pi_tag].append((date, summary, notes, consultant, client, float(hours), float(travel_hours), float(cumul_hours)))

        #
        # Calculate the number of free hours and billable hours in this transaction.
        #
        start_hours_used = float(cumul_hours) - float(hours) - float(travel_hours)

        if start_hours_used < CONSULTING_HOURS_FREE:
            free_hours_remaining = CONSULTING_HOURS_FREE - start_hours_used
        else:
            free_hours_remaining = 0

        if hours < free_hours_remaining:
            free_hours_used = hours
        else:
            free_hours_used = free_hours_remaining

        billable_hours = hours - free_hours_used + (travel_hours * CONSULTING_TRAVEL_RATE_DISCOUNT)

        # Save the consulting charges in a list of items for each PI.
        pi_tag_to_consulting_charges[pi_tag].append((date, summary, consultant, client, float(hours), float(travel_hours), float(billable_hours)))


# Generates the Billing sheet of a BillingNotifications (or BillingAggregate) workbook for a particular pi_tag.
# It uses dicts pi_tag_to_folder_sizes, and pi_tag_to_account_username_cpus, and puts summaries of its
# results in dict pi_tag_to_charges.
def generate_billing_sheet(wkbk, sheet, pi_tag, begin_month_timestamp, end_month_timestamp):

    global pi_tag_to_charges

    # Get affiliation status for the current PI.
    affiliation = pi_tag_to_affiliation[pi_tag]

    #
    # Set the column and row widths to contain all our data.
    #

    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # Give the first column 1 unit of space.
    col_dim_holder["A"] = ColumnDimension(sheet, index="A", width=1)
    # Give the second column 40 units of space.
    col_dim_holder["B"] = ColumnDimension(sheet, index="B", width=40)
    # Give the third, fourth, and fifth columns 11 units of space each.
    col_dim_holder["C"] = ColumnDimension(sheet, index="C", width=11)
    col_dim_holder["D"] = ColumnDimension(sheet, index="D", width=11)
    col_dim_holder["E"] = ColumnDimension(sheet, index="E", width=11)
    sheet.column_dimensions = col_dim_holder

    row_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # Give the first row 50 units of space.  ("Bill for Services Rendered")
    row_dim_holder[1] = RowDimension(sheet, index=1, ht=50)
    # Give the second row 30 units of space. ("PI: <PI NAME>")
    row_dim_holder[2] = RowDimension(sheet, index=2, ht=30)
    sheet.row_dimensions = row_dim_holder

    #
    # Write out the Document Header first ("Bill for Services Rendered")
    #

    # Write the text of the first row, with the GBSC address in merged columns.
    fmt = make_format(wkbk, {'font_size': 18, 'bold': True, 'underline': True,
                             'align': 'left', 'valign': 'vcenter'})
    sheet.cell(1, 2, 'Bill for Services Rendered').style = fmt

    fmt = make_format(wkbk, {'font_size': 12, 'text_wrap': True})
    sheet.merge_cells('C1:F1')
    sheet.cell(1, 3, "Genetics Bioinformatics Service Center (GBSC)\nSoM Technology & Innovation Center\n3165 Porter Drive, Palo Alto, CA").style = fmt

    # Write the PI name on the second row.
    (pi_first_name, pi_last_name, _) = pi_tag_to_names_email[pi_tag]

    fmt = make_format(wkbk, {'font_size' : 16, 'align': 'left', 'valign': 'vcenter'})
    sheet.cell(2, 2, "PI: %s, %s" % (pi_last_name, pi_first_name)).style = fmt

    #
    # Write the Billing Period dates on the fourth row.
    #
    begin_date_string = from_timestamp_to_date_string(begin_month_timestamp)

    # If we are running this script mid-month, use today's date as the end date for the Billing Period.
    now_timestamp = time.time()
    if now_timestamp < end_month_timestamp:
        end_date_string = from_timestamp_to_date_string(now_timestamp)
    else:
        end_date_string = from_timestamp_to_date_string(end_month_timestamp-1)

    billing_period_string = "Billing Period: %s - %s" % (begin_date_string, end_date_string)

    fmt = make_format(wkbk, { 'font_size': 14, 'bold': True})
    sheet.cell(4, 2, billing_period_string).style = fmt

    #
    # Calculate Breakdown of Charges first, then use those cumulative
    #  totals to fill out the Summary of Charges.
    #

    # Set up some formats for use in these tables.
    border_style = 'thin'

    # For "Summary of Charges" and "Breakdown of Charges"
    top_header_fmt = make_format(wkbk, {'font_size': 14, 'bold': True, 'underline': True,
                                        'align': 'left',
                                        'top': border_style, 'left': border_style})
    # For "Storage", "Computing", "Cloud Services", and "Bioinformatics Consulting" headers
    header_fmt = make_format(wkbk, {'font_size': 12, 'bold': True, 'underline': True,
                                    'align': 'left',
                                    'left': border_style})
    # Same as above, but with no underline: for the Summary table.
    header_no_ul_fmt = make_format(wkbk, {'font_size': 12, 'bold': True,
                                          'align': 'right',
                                          'left': border_style})

    # For subheaders within subtables, like "Account: XXX" in Computing subtable.
    sub_header_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                        'align': 'right', 'underline': True,
                                        'left': border_style})
    # For subheaders within subtables, but without underline.
    sub_header_no_ul_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                        'align': 'right',
                                        'left': border_style})

    # For column headers in subtables.
    col_header_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                        'align': 'right'})
    # As above, but allowing text wrap for long column headers (see Bioinformatics Computing subtable)
    col_header_textwrap_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                                 'align': 'right',
                                                 'text_wrap': True})
    # As sub_header_fmt, but with a border on the left.
    col_header_left_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                             'align': 'right',
                                             'left': border_style})
    # As sub_header_fmt, but with a border on the right.
    col_header_right_fmt = make_format(wkbk, {'font_size': 11, 'bold': True,
                                              'align': 'right',
                                              'right': border_style})

    # Text entry in a subtable (border on the left).
    item_entry_fmt = make_format(wkbk, {'font_size': 10,
                                        'align': 'right',
                                        'left': border_style})
    # As above, plus allowing text wrap (for descriptions in Bioinformatics Consulting).
    item_entry_textwrap_fmt = make_format(wkbk, {'font_size': 10,
                                                 'align': 'right',
                                                 'left': border_style,
                                                 'text_wrap': True})
    item_entry_italics_fmt = make_format(wkbk, {'font_size': 10,
                                                'align': 'right',
                                                'left': border_style,
                                                'italic': True})
    # Float entry in a subtable.
    float_entry_fmt = make_format(wkbk, {'font_size': 10,
                                         'align': 'right',
                                         'num_format': '0.0'})
    # As above, but vertically aligned to top (for Bioinformatics Consulting table).
    float_entry_valign_top_fmt = make_format(wkbk, {'font_size': 10,
                                                     'align': 'right',
                                                     'valign': 'top',
                                                     'num_format': '0.0'})
    # Integer entry in a subtable (not used?).
    int_entry_fmt = make_format(wkbk, {'font_size': 10,
                                       'align': 'right',
                                       'num_format': '0'})
    # Percentage entry in a subtable.
    pctage_entry_fmt = make_format(wkbk, {'font_size': 10,
                                          'num_format': '0%'})
    # String entry in a subtable, aligned right.
    string_entry_fmt = make_format(wkbk,{'font_size': 10,
                                         'align': 'right'})
    # As above, but vertically aligned to top (for Bioinformatics Consulting table).
    string_entry_valign_top_fmt = make_format(wkbk,{'font_size': 10,
                                                    'align': 'right',
                                                    'valign': 'top'})
    # Cost entry in Cloud Services subtable.
    cost_fmt = make_format(wkbk, {'font_size': 10,
                                  'align': 'right',
                                  'num_format': '$#,##0.00'})

    # Charge entry in subtables (with border on right).
    charge_fmt = make_format(wkbk, {'font_size': 10,
                                    'align': 'right',
                                    'right': border_style,
                                    'num_format': '$#,##0.00'})
    # As above, only vertically aligned to the top.
    charge_valign_top_fmt = make_format(wkbk, {'font_size': 10,
                                               'align': 'right',
                                               'valign': 'top',
                                               'right': border_style,
                                               'num_format': '$#,##0.00'})
    # Charge entry in Summary of Charges table.
    big_charge_fmt = make_format(wkbk, {'font_size': 12,
                                        'align': 'right',
                                        'right': border_style,
                                        'num_format': '$#,##0.00'})
    # As above, only bold (for grand total).
    big_bold_charge_fmt = make_format(wkbk, {'font_size': 12, 'bold': True,
                                             'align': 'right',
                                             'right': border_style, 'bottom': border_style,
                                             'num_format': '$#,##0.00'})
    # "Total XXX" entry for subtotals within subtables.
    bot_header_fmt = make_format(wkbk, {'font_size': 14, 'bold': True,
                                        'align': 'right',
                                        'left': border_style})
    # As above, but with a bottom border.
    bot_header_border_fmt = make_format(wkbk, {'font_size': 14, 'bold': True,
                                               'align': 'right',
                                               'left': border_style,
                                               'bottom': border_style})
    # Formats for borders in cells.
    upper_right_border_fmt = make_format(wkbk, {'top': border_style, 'right': border_style})
    lower_right_border_fmt = make_format(wkbk, {'bottom': border_style, 'right': border_style})
    lower_left_border_fmt  = make_format(wkbk, {'bottom': border_style, 'left': border_style})
    left_border_fmt = make_format(wkbk, {'left': border_style})
    right_border_fmt = make_format(wkbk, {'right': border_style})
    top_border_fmt = make_format(wkbk, {'top': border_style})
    bottom_border_fmt = make_format(wkbk, {'bottom': border_style})

    ######
    #
    # "Breakdown of Charges" (B14:??)
    #
    ######

    # Start the Breakdown of Charges table on the fifteenth row.
    curr_row = 15
    sheet.cell(curr_row, 2, "Breakdown of Charges:").style = top_header_fmt
    sheet.cell(curr_row, 3, None).style = top_border_fmt
    sheet.cell(curr_row, 4, None).style = top_border_fmt
    sheet.cell(curr_row, 5, None).style = upper_right_border_fmt

    curr_row += 1

    ###
    #
    # STORAGE Subtable of Breakdown of Charges table
    #
    ###

    # Skip line between "Breakdown of Charges".
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the "Storage" line.
    sheet.cell(curr_row, 2, "Storage:").style = header_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1

    total_storage_charges = 0.0
    total_storage_sizes   = 0.0

    # Get the rate from the Rates sheet of the BillingConfig workbook.
    cluster_acct_status = pi_tag_to_cluster_acct_status[pi_tag]
    if cluster_acct_status != "Full" and cluster_acct_status != "Free" and cluster_acct_status != "No":
        print("  Unexpected cluster status of '%s' for %s" % (cluster_acct_status, pi_tag), file=sys.stderr)

    storage_access_string = "%s Tier" % (cluster_acct_status.capitalize())

    (base_storage_rate, base_storage_rate_a1_cell) = (
        get_rate_amount_and_a1_cell_from_prefix(billing_config_wkbk,"Local HPC Storage", cluster_acct_status, "Base Storage", affiliation))
    (addl_storage_rate, addl_storage_rate_a1_cell) = (
        get_rate_amount_and_a1_cell_from_prefix(billing_config_wkbk,"Local HPC Storage", cluster_acct_status, "Additional Storage", affiliation))

    # Find lab folder in pi_tag_to_folder_sizes
    #  If found:
    #    Lab Folder
    #       Base Storage
    #       Additional Storage
    #       Total Storage - Lab Folder
    #    Other Folders
    #       folder 1
    #       folder 2
    #       Total Storage - Other Folders
    #    Total Storage
    #  If not:
    #    Folders
    #       folder 1
    #       folder 2
    #    Total Storage
    #

    # Find lab folder
    lab_folder_items = [item for item in pi_tag_to_folder_sizes[pi_tag] if item[0] == '/labs/%s' % pi_tag]

    # How many lab folders are there?  Hopefully, just one
    if len(lab_folder_items) == 1:

        # Get lab folder name, size, and percentage.
        (lab_folder_name, lab_folder_size, lab_folder_pctage) = lab_folder_items[0]

        # Write the storage headers.
        sheet.cell(curr_row, 2, "Lab Folder : %s" % lab_folder_name).style = sub_header_fmt
        sheet.cell(curr_row, 3, "Storage (Tb)").style = col_header_fmt
        sheet.cell(curr_row, 4, "%age").style = col_header_fmt
        sheet.cell(curr_row, 5, "Charge").style = col_header_right_fmt
        curr_row += 1

        starting_storage_row = curr_row

        if lab_folder_size >= BASE_STORAGE_SIZE:

            # Write the Base Storage line
            sheet.cell(curr_row, 2, "Base Storage").style = item_entry_fmt
            sheet.cell(curr_row, 3, BASE_STORAGE_SIZE).style = float_entry_fmt
            sheet.cell(curr_row, 4, lab_folder_pctage).style = pctage_entry_fmt
            pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
            sheet.cell(curr_row, 5, '=%s*%s' % (pctage_a1_cell, base_storage_rate_a1_cell)).style = charge_fmt

            ending_storage_row = curr_row

            total_storage_sizes   += BASE_STORAGE_SIZE
            total_storage_charges += base_storage_rate

            curr_row += 1

            # Check for additional storage for lab
            lab_folder_addl_storage = lab_folder_size - BASE_STORAGE_SIZE

        else:
            lab_folder_addl_storage = lab_folder_size

        if lab_folder_addl_storage > 0:

            # Write line with additional storage amount
            sheet.cell(curr_row, 2, "Additional Storage").style = item_entry_fmt
            sheet.cell(curr_row, 3, lab_folder_addl_storage).style = float_entry_fmt
            sheet.cell(curr_row, 4, lab_folder_pctage).style = pctage_entry_fmt
            cost_a1_cell = rowcol_to_a1_cell(curr_row, 3)
            pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
            sheet.cell(curr_row, 5, '=%s*%s*%s' % (cost_a1_cell, pctage_a1_cell, addl_storage_rate_a1_cell)).style = charge_fmt

            ending_storage_row = curr_row
            curr_row += 1

            total_storage_sizes   += lab_folder_addl_storage
            total_storage_charges += lab_folder_addl_storage * addl_storage_rate

        # Skip the line before Total Storage - "lab folder".
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1

        # Write Total Storage sum line for lab folder
        sheet.cell(curr_row, 2, "Total Storage - %s:" % lab_folder_name).style = sub_header_no_ul_fmt

        top_storage_charges_a1_cell = rowcol_to_a1_cell(starting_storage_row, 3)
        bot_storage_charges_a1_cell = rowcol_to_a1_cell(ending_storage_row + 1, 3)
        sheet.cell(curr_row, 3,
            '=SUM(%s:%s)' % (top_storage_charges_a1_cell, bot_storage_charges_a1_cell)).style = float_entry_fmt
        # Nothing in pctage cell (col 4)
        top_storage_charges_a1_cell = rowcol_to_a1_cell(starting_storage_row, 5)
        bot_storage_charges_a1_cell = rowcol_to_a1_cell(ending_storage_row + 1, 5)
        sheet.cell(curr_row, 5,
                   '=SUM(%s:%s)' % (top_storage_charges_a1_cell, bot_storage_charges_a1_cell)).style = charge_fmt

        lab_folder_total_sizes_a1_cell   = rowcol_to_a1_cell(curr_row, 3)  # For sum of Total Storage formula
        lab_folder_total_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)
        curr_row += 1

        # Remove the lab folder from the pi_tag_to_folder_sizes list
        pi_tag_to_folder_sizes[pi_tag].remove(lab_folder_items[0])

    else:
        lab_folder_total_sizes_a1_cell   = None
        lab_folder_total_charges_a1_cell = None

    # Are there more folders to list?
    if len(pi_tag_to_folder_sizes[pi_tag]) > 0:

        other_folders_storage_sizes = 0.0

        # Skip row after lab folder section
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1 # Skip row after first lab folder section

        sheet.cell(curr_row, 2, "Other Folders").style = sub_header_fmt
        # Nothing in other columns
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1

        starting_storage_row = curr_row

        for (folder, size, pctage) in pi_tag_to_folder_sizes[pi_tag]:
            sheet.cell(curr_row, 2, folder).style = item_entry_fmt
            sheet.cell(curr_row, 3, size).style = float_entry_fmt
            sheet.cell(curr_row, 4, pctage).style = pctage_entry_fmt

            # Calculate charges.
            if addl_storage_rate is not None:
                charge = size * pctage * addl_storage_rate
                total_storage_charges += charge
            else:
                charge = "No rate"

            total_storage_sizes += size
            other_folders_storage_sizes += size

            cost_a1_cell = rowcol_to_a1_cell(curr_row, 3)
            pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
            sheet.cell(curr_row, 5,
                       '=%s*%s*%s' % (cost_a1_cell, pctage_a1_cell, addl_storage_rate_a1_cell)).style = charge_fmt

            # Keep track of last row with storage values.
            ending_storage_row = curr_row

            # Advance to the next row.
            curr_row += 1

        # Skip row after other folder section
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1  # Skip row

        # Write Total Storage sum line for lab folder
        sheet.cell(curr_row, 2, "Total Storage - Other Folders:").style = sub_header_no_ul_fmt

        top_storage_charges_a1_cell = rowcol_to_a1_cell(starting_storage_row, 3)
        bot_storage_charges_a1_cell = rowcol_to_a1_cell(ending_storage_row + 1, 3)
        sheet.cell(curr_row, 3,
                   '=SUM(%s:%s)' % (top_storage_charges_a1_cell, bot_storage_charges_a1_cell)).style = float_entry_fmt

        # Nothing in pctage cell (col 4)

        top_storage_charges_a1_cell = rowcol_to_a1_cell(starting_storage_row, 5)
        bot_storage_charges_a1_cell = rowcol_to_a1_cell(ending_storage_row + 1, 5)
        sheet.cell(curr_row, 5,
                   '=SUM(%s:%s)' % (top_storage_charges_a1_cell, bot_storage_charges_a1_cell)).style = charge_fmt

        other_folders_total_sizes_a1_cell = rowcol_to_a1_cell(curr_row, 3)  # For sum of Total Storage formula
        other_folders_total_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)

        curr_row += 1
    else:
        other_folders_total_sizes_a1_cell = None
        other_folders_total_charges_a1_cell = None

    # Skip the line before Total Storage.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1

    # Write the Total Storage line.
    sheet.cell(curr_row, 2, "Total Storage:").style = bot_header_fmt
    if lab_folder_total_sizes_a1_cell is not None and other_folders_total_sizes_a1_cell is not None:
        sheet.cell(curr_row, 3, '=SUM(%s,%s)' % (lab_folder_total_sizes_a1_cell, other_folders_total_sizes_a1_cell)).style = float_entry_fmt
    elif lab_folder_total_sizes_a1_cell is not None:
        sheet.cell(curr_row, 3, '=%s' % lab_folder_total_sizes_a1_cell).style = float_entry_fmt
    elif other_folders_total_sizes_a1_cell is not None:
        sheet.cell(curr_row, 3, '=%s' % other_folders_total_sizes_a1_cell).style = float_entry_fmt
    else:
        sheet.cell(curr_row, 3, '').style = float_entry_fmt
    # Nothing in pctage column (col 4)
    if lab_folder_total_charges_a1_cell is not None and other_folders_total_charges_a1_cell is not None:
        sheet.cell(curr_row, 5, '=SUM(%s,%s)' % (lab_folder_total_charges_a1_cell, other_folders_total_charges_a1_cell)).style = charge_fmt
    elif lab_folder_total_charges_a1_cell is not None:
        sheet.cell(curr_row, 5, '=%s' % lab_folder_total_charges_a1_cell).style = charge_fmt
    elif other_folders_total_charges_a1_cell is not None:
        sheet.cell(curr_row, 5, '=%s' % other_folders_total_charges_a1_cell).style = charge_fmt
    else:
        sheet.cell(curr_row, 5, '').style = charge_fmt

    # Save reference to this cell for use in Summary subtable.
    total_storage_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)

    curr_row += 1

    # Skip the next line and draw line under this row.
    sheet.cell(curr_row, 2, None).style = lower_left_border_fmt
    sheet.cell(curr_row, 3, None).style = bottom_border_fmt
    sheet.cell(curr_row, 4, None).style = bottom_border_fmt
    sheet.cell(curr_row, 5, None).style = lower_right_border_fmt
    curr_row += 1

    ###
    #
    # COMPUTING Subtable of Breakdown of Charges table
    #
    ###

    # Get the rate from the Rates sheet of the BillingConfig workbook.
    computing_access_string = "%s Tier" % (cluster_acct_status.capitalize())

    # Get both rates for CPU, in case someone outside the lab runs a job for a Free Tier lab (usually Consulting).
    (free_tier_cpu_rate, free_tier_cpu_rate_a1_cell) = \
        get_rate_amount_and_a1_cell_from_prefix(billing_config_wkbk,"Local Computing", "Free", None, affiliation)
    (full_tier_cpu_rate, full_tier_cpu_rate_a1_cell) = \
        get_rate_amount_and_a1_cell_from_prefix(billing_config_wkbk,"Local Computing", "Full", None, affiliation)

    # Choose the default rate for the lab.
    if cluster_acct_status != "Free":
        (cpu_rate, cpu_rate_a1_cell) = (full_tier_cpu_rate, full_tier_cpu_rate_a1_cell)
    else:
        (cpu_rate, cpu_rate_a1_cell) = (free_tier_cpu_rate, free_tier_cpu_rate_a1_cell)

    # Skip row before Computing header.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the Computing line.
    sheet.cell(curr_row, 2, "Computing:").style = header_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1

    # Loop over pi_tag_to_account_username_cpus for account/username combos.
    account_username_cpus_list = pi_tag_to_account_username_cpus.get(pi_tag)

    # The list of "Total Charges" rows for each account.
    total_computing_charges_row_list = []

    total_computing_charges = 0.0
    total_computing_cpuhrs = 0.0

    if account_username_cpus_list is not None:

        for (account, username_cpu_pctage_list) in account_username_cpus_list:

            # Write the account subheader.
            if account != "":
                sheet.cell(curr_row, 2, "Account: %s" % account).style = sub_header_fmt
            else:
                sheet.cell(curr_row, 2, "Account: Lab Default").style = sub_header_fmt
            # sheet.cell(curr_row, 4, None).style = col_header_right_fmt
            sheet.cell(curr_row, 5, None).style = col_header_right_fmt
            curr_row += 1

            # Skip row after account subheader.
            sheet.cell(curr_row, 2, None).style = left_border_fmt
            sheet.cell(curr_row, 5, None).style = right_border_fmt
            curr_row += 1

            # Write the computing headers.
            sheet.cell(curr_row, 2, "User").style = col_header_left_fmt
            sheet.cell(curr_row, 3, "CPU units").style = col_header_fmt
            sheet.cell(curr_row, 4, "%age").style = col_header_fmt
            sheet.cell(curr_row, 5, "Charge").style = col_header_right_fmt
            curr_row += 1

            # Get the job details for the users associated with this PI.
            starting_computing_row = curr_row
            ending_computing_row   = curr_row
            if len(username_cpu_pctage_list) > 0:

                for (username, cpu_units, pctage) in username_cpu_pctage_list:

                    pi_tags_for_username = get_pi_tags_for_username_by_date(username, begin_month_timestamp)

                    if pi_tag in [pi_pct[0] for pi_pct in pi_tags_for_username]:
                        username_fmt = item_entry_fmt
                        user_cpu_rate = cpu_rate
                        user_cpu_rate_a1_cell  = cpu_rate_a1_cell
                    else:
                        username_fmt = item_entry_italics_fmt
                        user_cpu_rate = full_tier_cpu_rate
                        user_cpu_rate_a1_cell  = full_tier_cpu_rate_a1_cell

                    fullname = username_to_user_details[username][1]
                    sheet.cell(curr_row, 2, "%s (%s)" % (fullname, username)).style = username_fmt
                    sheet.cell(curr_row, 3, cpu_units).style = float_entry_fmt
                    sheet.cell(curr_row, 4, pctage).style = pctage_entry_fmt

                    if user_cpu_rate is not None:
                        charge = cpu_units * pctage * user_cpu_rate

                        # Check if user has accumulated more than $5000 in a month.
                        if charge > 5000:
                            print("  *** User %s (%s) for PI %s, Account %s: $%0.02f" % (username_to_user_details[username][1], username, pi_tag, account, charge))

                        total_computing_charges += charge

                    total_computing_cpuhrs += cpu_units

                    cpu_a1_cell    = rowcol_to_a1_cell(curr_row, 3)
                    pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
                    sheet.cell(curr_row, 5, '=%s*%s*%s' % (cpu_a1_cell, pctage_a1_cell, user_cpu_rate_a1_cell)).style = charge_fmt

                    # Keep track of last row with computing values.
                    ending_computing_row = curr_row

                    # Advance to the next row.
                    curr_row += 1

                # Skip row after last user.
                sheet.cell(curr_row, 2, None).style = left_border_fmt
                sheet.cell(curr_row, 5, None).style = right_border_fmt
                curr_row += 1

            else:
                # No users for this PI.
                sheet.cell(curr_row, 2, "No jobs").style = item_entry_fmt
                sheet.cell(curr_row, 5, 0.0).style = charge_fmt
                curr_row += 1

            # Write the Total Charges line header.
            if account != "":
                sheet.cell(curr_row, 2, "Total charges - %s:" % account).style = col_header_left_fmt
            else:
                sheet.cell(curr_row, 2, "Total charges - Lab Default:").style = col_header_left_fmt

            # Write the formula for the CPU-core-hrs subtotal for the account.
            top_cpu_a1_cell = rowcol_to_a1_cell(starting_computing_row, 3)
            bot_cpu_a1_cell = rowcol_to_a1_cell(ending_computing_row, 3)
            sheet.cell(curr_row, 3, '=SUM(%s:%s)' % (top_cpu_a1_cell, bot_cpu_a1_cell)).style = float_entry_fmt

            sheet.cell(curr_row, 4, None).style = col_header_fmt

            # Write the formula for the charges subtotal for the account.
            top_charge_a1_cell = rowcol_to_a1_cell(starting_computing_row, 5)
            bot_charge_a1_cell = rowcol_to_a1_cell(ending_computing_row + 1, 5)
            sheet.cell(curr_row, 5, '=SUM(%s:%s)' % (top_charge_a1_cell, bot_charge_a1_cell)).style = charge_fmt

            # Save row of this total charges for the account for Total Computing charges sum.
            total_computing_charges_row_list.append(curr_row)

            curr_row += 1

            # Skip row after account subtotal.
            sheet.cell(curr_row, 2, None).style = left_border_fmt
            sheet.cell(curr_row, 5, None).style = right_border_fmt
            curr_row += 1

    # Write the Total Computing line.
    sheet.cell(curr_row, 2, "Total Computing:").style = bot_header_fmt
    # sheet.cell(curr_row, 3, total_computing_cpuhrs).style = float_entry_fmt
    sheet.cell(curr_row, 5, total_computing_charges).style = charge_fmt

    if len(total_computing_charges_row_list) > 0:

        total_cpu_cell_list = [rowcol_to_a1_cell(x, 3) for x in total_computing_charges_row_list]
        total_computing_charges_cell_list = [rowcol_to_a1_cell(x, 5) for x in total_computing_charges_row_list]

        # Create formula from account total CPU cells.
        total_cpu_formula = "=" + "+".join(total_cpu_cell_list)
        sheet.cell(curr_row, 3, total_cpu_formula).style = float_entry_fmt

        # Create formula from account total charges cells.
        total_computing_charges_formula = "=" + "+".join(total_computing_charges_cell_list)

        # sheet.write_formula(curr_row, 4, total_computing_charges_formula, charge_fmt)
        sheet.cell(curr_row, 5, total_computing_charges_formula).style = charge_fmt

    else:
        sheet.cell(curr_row, 5, 0.0).style = charge_fmt

    # Save reference to this cell for use in Summary subtable.
    total_computing_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)

    curr_row += 1

    # Skip the next line and draw line under this row.
    sheet.cell(curr_row, 2, None).style = lower_left_border_fmt
    sheet.cell(curr_row, 3, None).style = bottom_border_fmt
    sheet.cell(curr_row, 4, None).style = bottom_border_fmt
    sheet.cell(curr_row, 5, None).style = lower_right_border_fmt
    curr_row += 1

    ###
    #
    # CLOUD SERVICES Subtable of Breakdown of Charges table
    #
    ###

    # Skip line between previous subtable.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the "Cloud Services" line.
    sheet.cell(curr_row, 2, "Cloud Services:").style = header_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1

    # Get the rate from the Rates sheet of the BillingConfig workbook.
    rate_cloud_per_dollar = get_rate(billing_config_wkbk, 'Cloud Services - %s' % affiliation)
    rate_cloud_a1_cell    = get_rate_a1_cell(billing_config_wkbk, 'Cloud Services - %s' % affiliation)

    total_cloud_charges = 0.0

    # The list of "Total Charges" rows for each account.
    total_cloud_charges_row_list = []

    # For all the cloud accounts for this PI:
    pi_cloud_account_pctages = pi_tag_to_cloud_account_pctages[pi_tag]

    for (account, pctage) in pi_cloud_account_pctages:

        account_name = cloud_account_to_account_names[account]

        # Write the account subheader.
        if account_name is not None and account_name != "":
            sheet.cell(curr_row, 2, "Account: %s" % account_name).style = sub_header_fmt
        else:
            sheet.cell(curr_row, 2, "Account: %s" % account).style = sub_header_fmt
        sheet.cell(curr_row, 5, None).style = col_header_right_fmt
        curr_row += 1

        # Skip row after account subheader.
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1

        # Write the cloud services headers.
        sheet.cell(curr_row, 2, "Project").style = col_header_left_fmt
        sheet.cell(curr_row, 3, "Cost").style = col_header_fmt
        sheet.cell(curr_row, 4, "%age").style = col_header_fmt
        sheet.cell(curr_row, 5, "Charge").style = col_header_right_fmt
        curr_row += 1

        total_cloud_account_charges = 0.0

        starting_cloud_row = curr_row
        ending_cloud_row   = curr_row - 1 # Inverted order of start and end means "no projects found".

        for project in cloud_account_to_cloud_projects[account]:

            project_cost = cloud_project_account_to_total_charges[(project, account)]

            if project_cost != 0.0:
                # A blank project name means (usually) a credit applied to the account.
                if project is not None:
                    # If we have the project number here, use the project name.
                    if project[0].isdigit():
                        sheet.cell(curr_row, 2, cloud_projnum_to_cloud_project[project]).style = item_entry_fmt
                    else:
                        sheet.cell(curr_row, 2, project).style = item_entry_fmt
                else:
                    sheet.cell(curr_row, 2, "Misc charges/credits").style = item_entry_fmt
                sheet.cell(curr_row, 3, project_cost).style = cost_fmt
                sheet.cell(curr_row, 4, pctage).style = pctage_entry_fmt

                # Calculate charges.
                charge = project_cost * pctage * rate_cloud_per_dollar
                total_cloud_account_charges += charge

                # Write formula for charges to the sheet.
                cost_a1_cell   = rowcol_to_a1_cell(curr_row, 3)
                pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
                sheet.cell(curr_row, 5, '=%s*%s*%s' % (cost_a1_cell, pctage_a1_cell, rate_cloud_a1_cell)).style = charge_fmt

                # Keep track of last row with cloud project values.
                ending_cloud_row = curr_row

                # Advance to the next row.
                curr_row += 1

        total_cloud_charges += total_cloud_account_charges

        # If there were no projects, put a row saying so.
        if starting_cloud_row > ending_cloud_row:
            sheet.cell(curr_row, 2, "No Projects").style = item_entry_fmt

            cost_a1_cell = rowcol_to_a1_cell(curr_row, 3)
            pctage_a1_cell = rowcol_to_a1_cell(curr_row, 4)
            sheet.cell(curr_row, 5, "=%s*%s*%s" % (cost_a1_cell, pctage_a1_cell, rate_cloud_a1_cell)).style = charge_fmt

            curr_row += 1
            ending_cloud_row = starting_cloud_row

        # Skip the line before "Total charges - ACCOUNT".
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1

        # Write the Total Charges line header.
        if account_name is not None and account_name != "":
            sheet.cell(curr_row, 2, "Total charges - %s:" % account_name).style = col_header_left_fmt
        else:
            sheet.cell(curr_row, 2, "Total charges - %s:" % account).style = col_header_left_fmt

        # Write the formula for the charges subtotal for the account.
        top_charge_a1_cell = rowcol_to_a1_cell(starting_cloud_row, 5)
        bot_charge_a1_cell = rowcol_to_a1_cell(ending_cloud_row + 1, 5)
        sheet.cell(curr_row, 5, '=SUM(%s:%s)' % (top_charge_a1_cell, bot_charge_a1_cell)).style = charge_fmt

        # Save row of this total charges for the account for Total Cloud charges sum.
        total_cloud_charges_row_list.append(curr_row)

        curr_row += 1

        # Skip row after account subtotal.
        sheet.cell(curr_row, 2, None).style = left_border_fmt
        sheet.cell(curr_row, 5, None).style = right_border_fmt
        curr_row += 1

    # Skip the line before "Total Cloud Services".
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1

    # Write the "Total Cloud Services" line.
    sheet.cell(curr_row, 2, "Total Cloud Services:").style = bot_header_fmt

    if len(total_cloud_charges_row_list) > 0:

        total_cloud_charges_cell_list = [rowcol_to_a1_cell(x, 5) for x in total_cloud_charges_row_list]

        # Create formula from account total charges cells.
        total_cloud_charges_formula = "=" + "+".join(total_cloud_charges_cell_list)

        # sheet.write_formula(curr_row, 4, total_computing_charges_formula, charge_fmt)
        sheet.cell(curr_row, 5, total_cloud_charges_formula).style = charge_fmt

    else:
        sheet.cell(curr_row, 5, 0.0).style = charge_fmt

    # Save reference to this cell for use in Summary subtable.
    total_cloud_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)

    curr_row += 1

    # Skip the next line and draw line under this row.
    sheet.cell(curr_row, 2, None).style = lower_left_border_fmt
    sheet.cell(curr_row, 3, None).style = bottom_border_fmt
    sheet.cell(curr_row, 4, None).style = bottom_border_fmt
    sheet.cell(curr_row, 5, None).style = lower_right_border_fmt
    curr_row += 1

    ###
    #
    # CONSULTING Subtable of Breakdown of Charges table
    #
    ###

    # Skip row before Bioinformatics Consulting header.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the Bioinformatics Consulting line.
    sheet.cell(curr_row, 2, "Bioinformatics Consulting (BaaS):").style = header_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the consulting headers.
    sheet.cell(curr_row, 2, "Date: Task (Consultant) [Client]").style = col_header_left_fmt
    sheet.cell(curr_row, 3, "Hours (Travel Hours)").style = col_header_textwrap_fmt
    sheet.cell(curr_row, 4, "Billable Hours").style = col_header_textwrap_fmt
    sheet.cell(curr_row, 5, "Charge").style = col_header_right_fmt
    curr_row += 1

    total_consulting_hours = 0.0
    total_consulting_travel_hours = 0.0
    total_consulting_billable_hours = 0.0
    total_consulting_charges = 0.0

    starting_consulting_row = curr_row

    # Get the rate from the Rates sheet of the BillingConfig workbook.
    rate_consulting_per_hour = get_rate(billing_config_wkbk, 'Bioinformatics Consulting - %s' % affiliation)
    rate_consulting_a1_cell = get_rate_a1_cell(billing_config_wkbk, 'Bioinformatics Consulting - %s' % affiliation)

    if len(pi_tag_to_consulting_details[pi_tag]) > 0:

        for (date, summary, consultant, client, hours, travel_hours, billable_hours) in pi_tag_to_consulting_charges[pi_tag]:

            date_task_consultant_str = "%s: %s (%s) [%s]" % (from_datetime_to_date_string(date), summary, consultant, client)
            sheet.cell(curr_row, 2, date_task_consultant_str).style = item_entry_textwrap_fmt

            hours_travel_hours_str = "%s (%s)" % (hours, travel_hours)
            sheet.cell(curr_row, 3, hours_travel_hours_str).style = string_entry_valign_top_fmt
            sheet.cell(curr_row, 4, billable_hours).style = float_entry_valign_top_fmt

            charge = rate_consulting_per_hour * billable_hours
            total_consulting_charges += charge

            total_consulting_billable_hours += billable_hours
            total_consulting_hours += hours
            total_consulting_travel_hours += travel_hours

            billable_hours_a1_cell = rowcol_to_a1_cell(curr_row, 4)
            sheet.cell(curr_row, 5, '=%s*%s' % (billable_hours_a1_cell, rate_consulting_a1_cell)).style = charge_valign_top_fmt
            curr_row += 1

    else:
        sheet.cell(curr_row, 2, "No consulting").style = item_entry_fmt

        billable_hours_a1_cell = rowcol_to_a1_cell(curr_row, 4)
        sheet.cell(curr_row, 5, '=%s*%s' % (billable_hours_a1_cell, rate_consulting_a1_cell)).style = charge_fmt
        curr_row += 1

    ending_consulting_row = curr_row

    # Skip the line before Total Consulting.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the Total Consulting line.
    sheet.cell(curr_row, 2, "Total Consulting:").style = bot_header_fmt
    sheet.cell(curr_row, 3, "%s (%s)" % (total_consulting_hours, total_consulting_travel_hours)).style = string_entry_fmt
    top_storage_charges_a1_cell = rowcol_to_a1_cell(starting_consulting_row, 4)
    bot_billable_hours_a1_cell = rowcol_to_a1_cell(ending_consulting_row, 4)
    sheet.cell(curr_row, 4, '=SUM(%s:%s)' % (top_storage_charges_a1_cell, bot_billable_hours_a1_cell)).style = float_entry_fmt
    top_charges_a1_cell = rowcol_to_a1_cell(starting_consulting_row, 5)
    bot_charges_a1_cell = rowcol_to_a1_cell(ending_consulting_row, 5)
    sheet.cell(curr_row, 5, '=SUM(%s:%s)' % (top_charges_a1_cell, bot_charges_a1_cell)).style = charge_fmt

    # Save reference to this cell for use in Summary subtable.
    total_consulting_charges_a1_cell = rowcol_to_a1_cell(curr_row, 5)

    curr_row += 1

    # Skip the next line and draw line under this row.
    sheet.cell(curr_row, 2, None).style = lower_left_border_fmt
    sheet.cell(curr_row, 3, None).style = bottom_border_fmt
    sheet.cell(curr_row, 4, None).style = bottom_border_fmt
    sheet.cell(curr_row, 5, None).style = lower_right_border_fmt
    curr_row += 1

    #####
    #
    # Summary of Charges table (B6:E12)
    #
    #####

    # Start the Summary of Charges table on the sixth row.
    curr_row = 6
    sheet.cell(curr_row, 2, "Summary of Charges:").style = top_header_fmt
    sheet.cell(curr_row, 3, None).style = top_border_fmt
    sheet.cell(curr_row, 4, None).style = top_border_fmt
    sheet.cell(curr_row, 5, None).style = upper_right_border_fmt
    curr_row += 1
    # Write the Storage line.
    sheet.cell(curr_row, 2, "Storage").style = header_no_ul_fmt
    sheet.cell(curr_row, 3, total_storage_sizes).style = float_entry_fmt
    sheet.cell(curr_row, 4, storage_access_string)
    sheet.cell(curr_row, 5, '=%s' % total_storage_charges_a1_cell).style = big_charge_fmt
    curr_row += 1
    # Write the Computing line.
    sheet.cell(curr_row, 2, "Computing").style = header_no_ul_fmt
    # sheet.cell(curr_row, 3, total_computing_cpuhrs).style = float_entry_fmt
    sheet.cell(curr_row, 4, computing_access_string)
    sheet.cell(curr_row, 5, '=%s' % total_computing_charges_a1_cell).style = big_charge_fmt
    curr_row += 1
    # Write the Cloud Services line.
    sheet.cell(curr_row, 2, "Cloud Services").style = header_no_ul_fmt
    sheet.cell(curr_row, 5, '=%s' % total_cloud_charges_a1_cell).style = big_charge_fmt
    curr_row += 1
    # Write the Consulting line.
    sheet.cell(curr_row, 2, "Bioinformatics Consulting").style = header_no_ul_fmt
    sheet.cell(curr_row, 3, total_consulting_billable_hours).style = float_entry_fmt
    sheet.cell(curr_row, 5, '=%s' % total_consulting_charges_a1_cell).style = big_charge_fmt
    curr_row += 1
    # Skip a line.
    sheet.cell(curr_row, 2, None).style = left_border_fmt
    sheet.cell(curr_row, 5, None).style = right_border_fmt
    curr_row += 1
    # Write the Grand Total line.
    sheet.cell(curr_row, 2, "Total Charges").style = bot_header_border_fmt
    sheet.cell(curr_row, 3, None).style = bottom_border_fmt
    sheet.cell(curr_row, 4, None).style = bottom_border_fmt
    total_charges = total_storage_charges + total_computing_charges + total_cloud_charges + total_consulting_charges
    sheet.cell(curr_row, 5, '=%s+%s+%s+%s' % (total_storage_charges_a1_cell, total_computing_charges_a1_cell, total_cloud_charges_a1_cell, total_consulting_charges_a1_cell)).style = big_bold_charge_fmt
    curr_row += 1

    #
    # Fill in row in pi_tag -> charges hash.

    pi_tag_to_charges[pi_tag] = [total_storage_charges, total_computing_charges, total_cloud_charges,
                                 total_consulting_charges,
                                 total_charges]

    # CHECK: If "Free Tier" and total_storage_charges >= 7 TB: flag an error
    if cluster_acct_status == "Free" and total_storage_charges >= 7:
        print("   *** Free Tier PI tag", pi_tag, "has", total_storage_charges, "TB", file=sys.stderr)


# Copies the Rates sheet from the Rates sheet in the BillingConfig workbook to
# a BillingNotification workbook.
def generate_rates_sheet(rates_input_sheet, pi_tag, rates_output_sheet):

    # Freeze the first row.
    rates_output_sheet.freeze_panes = 'A2'

    # Set the column widths
    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(rates_output_sheet)
    # "Type"
    col_dim_holder["A"] = ColumnDimension(rates_output_sheet, index="A", width=45)
    # "Amount"
    col_dim_holder["B"] = ColumnDimension(rates_output_sheet, index="B", width=8)
    # "Unit"
    col_dim_holder["C"] = ColumnDimension(rates_output_sheet, index="C", width=8)
    # "Time"
    col_dim_holder["D"] = ColumnDimension(rates_output_sheet, index="D", width=6)
    # "iLab Service ID"
    col_dim_holder["E"] = ColumnDimension(rates_output_sheet, index="E", width=12)
    rates_output_sheet.column_dimensions = col_dim_holder

    # Get the affliation and cluster status for the PI
    affiliation = pi_tag_to_affiliation[pi_tag]
    cluster_acct_status = pi_tag_to_cluster_acct_status[pi_tag]

    # Just copy the Rates sheet from the BillingConfig to the BillingNotification.
    curr_row = 2
    for row in rates_input_sheet.iter_rows(min_row=2, values_only=True):

        # If this row pertains to the PI's affiliation or cluster status, make the row bold.
        highlight_row = row[0] is not None and (affiliation in row[0] and ("Local" not in row[0] or cluster_acct_status in row[0]))

        # Write each value from row into output row of output Rates sheet.
        curr_col = 1
        for val in row:

            if curr_row == 1:
                rates_output_sheet.cell(curr_row, curr_col, val).style = BOLD_FORMAT
            elif curr_col == 2:
                if highlight_row:
                    rates_output_sheet.cell(curr_row, curr_col, val).style = BOLD_MONEY_FORMAT
                else:
                    rates_output_sheet.cell(curr_row, curr_col, val).style = MONEY_FORMAT
            elif highlight_row:
                rates_output_sheet.cell(curr_row, curr_col, val).style = BOLD_FORMAT
            else:
                rates_output_sheet.cell(curr_row, curr_col, val)
            curr_col += 1
        curr_row += 1


# Generates a Computing Details sheet for a BillingNotification workbook with
# job details associated with a particular PI.  It reads from dict pi_tag_to_job_details.
def generate_computing_details_sheet(wkbk, sheet, pi_tag):

    global BOLD_FORMAT

    # Freeze the first row.
    sheet.freeze_panes = 'A2'

    # Set the column widths
    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # "Job Date"
    col_dim_holder["A"] = ColumnDimension(sheet, index="A", width=10)
    # "Username"
    col_dim_holder["B"] = ColumnDimension(sheet, index="B", width=8)
    # "Job Name"
    col_dim_holder["C"] = ColumnDimension(sheet, index="C", width=40)
    # "Job Tag"
    col_dim_holder["D"] = ColumnDimension(sheet, index="D", width=14)
    # "Node"
    col_dim_holder["E"] = ColumnDimension(sheet, index="E", width=22)
    # "CPU-core Hours"
    col_dim_holder["F"] = ColumnDimension(sheet, index="F", width=8)
    # "Job ID"
    col_dim_holder["G"] = ColumnDimension(sheet, index="G", width=10)
    # "%age"
    col_dim_holder["H"] = ColumnDimension(sheet, index="H", width=6)
    sheet.column_dimensions = col_dim_holder

    # Count the number of sheets these detail lines go into
    sheet_count = 1

    # Write the job details, sorted by username.
    curr_row = 2
    for (date, username, job_name, account, node, cpu_core_hrs, jobID, pctage) in sorted(pi_tag_to_job_details[pi_tag],key=lambda s: s[1]):

        curr_col = 1
        sheet.cell(curr_row, curr_col, date).style = DATE_FORMAT ; curr_col += 1
        sheet.cell(curr_row, curr_col, username) ; curr_col += 1
        sheet.cell(curr_row, curr_col, job_name) ; curr_col += 1
        sheet.cell(curr_row, curr_col, account) ; curr_col += 1
        sheet.cell(curr_row, curr_col, node) ; curr_col += 1
        sheet.cell(curr_row, curr_col, cpu_core_hrs).style = FLOAT_FORMAT ; curr_col += 1
        sheet.cell(curr_row, curr_col, jobID) ; curr_col += 1
        sheet.cell(curr_row, curr_col, pctage).style = PERCENT_FORMAT ; curr_col += 1

        # Advance to the next row.
        curr_row += 1

        # If this sheet is full...
        if curr_row > EXCEL_MAX_ROWS:
            #
            # Create a new sheet
            #

            # Advance the sheet count.
            sheet_count += 1
            sheet_name = 'Computing Details {}'.format(sheet_count)
            sheet = wkbk.create_sheet(sheet_name)

            # Initialize the header line for the new sheet
            for col in range(0, len(BILLING_NOTIFS_SHEET_COLUMNS["Computing Details"])):
                sheet.cell(1, col + 1, BILLING_NOTIFS_SHEET_COLUMNS["Computing Details"][col]).style = BOLD_FORMAT

            # Freeze the first row.
            sheet.freeze_panes = 'A2'
            # Set the column dimensions.
            sheet.column_dimensions = col_dim_holder

            # Set the new next row to be the one after the header.
            curr_row = 2


# Generates the Lab Users sheet for a BillingNotification workbook with
# username details for a particular PI.  It reads from dicts:
#  cloud_project_account_to_cloud_details
#  pi_tag_to_cloud_account_pctages
def generate_cloud_details_sheet(sheet, pi_tag):

    # Freeze the first row.
    sheet.freeze_panes = 'A2'

    # Set the column widths
    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # "Platform"
    col_dim_holder["A"] = ColumnDimension(sheet, index="A", width=20)
    # "Project"
    col_dim_holder["B"] = ColumnDimension(sheet, index="B", width=25)
    # "Description"
    col_dim_holder["C"] = ColumnDimension(sheet, index="C", width=60)
    # "Dates"
    col_dim_holder["D"] = ColumnDimension(sheet, index="D", width=20)
    # "Quantity"
    col_dim_holder["E"] = ColumnDimension(sheet, index="E", width=12)
    # "Unit of Measure"
    col_dim_holder["F"] = ColumnDimension(sheet, index="F", width=25)
    # "Charge"
    col_dim_holder["G"] = ColumnDimension(sheet, index="G", width=10)
    # "%age"
    col_dim_holder["H"] = ColumnDimension(sheet, index="H", width=6)
    # "Cost"
    col_dim_holder["I"] = ColumnDimension(sheet, index="I", width=10)
    sheet.column_dimensions = col_dim_holder

    curr_row = 2
    
    # Get the list of accounts associated with this PI.
    for (account, pctage) in pi_tag_to_cloud_account_pctages[pi_tag]:

        for project in cloud_account_to_cloud_projects[account]:

            # Write the cloud details.
            for (platform, description, dates, quantity, uom, charge) in cloud_project_account_to_cloud_details[(project, account)]:

                curr_col = 1
                sheet.cell(curr_row, curr_col, platform);    curr_col += 1
                # If we have the project number here, use the project name.
                if project is not None and project[0].isdigit():
                    sheet.cell(curr_row, curr_col, cloud_projnum_to_cloud_project[project]);  curr_col += 1
                else:
                    sheet.cell(curr_row, curr_col, project) ; curr_col += 1
                sheet.cell(curr_row, curr_col, description); curr_col += 1
                sheet.cell(curr_row, curr_col, dates);       curr_col += 1
                sheet.cell(curr_row, curr_col, quantity).style = FLOAT_FORMAT;  curr_col += 1
                sheet.cell(curr_row, curr_col, uom);         curr_col += 1
                sheet.cell(curr_row, curr_col, charge).style = MONEY_FORMAT;    curr_col += 1
                sheet.cell(curr_row, curr_col, pctage).style = PERCENT_FORMAT;  curr_col += 1

                lab_cost = charge * pctage
                sheet.cell(curr_row, curr_col, lab_cost).style = MONEY_FORMAT; curr_col += 1

                # Advance to the next row.
                curr_row += 1


# Generates the Consulting Details sheet for a BillingNotifications workbook with
# consulting details for a particular PI.  It reads from dict:
#  pi_tag_to_consulting_details
def generate_consulting_details_sheet(sheet, pi_tag):

    # Freeze the first row.
    sheet.freeze_panes = 'A2'

    # Set the column widths
    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # "Date"
    col_dim_holder["A"] = ColumnDimension(sheet, index="A", width=9)
    # "Summary"
    col_dim_holder["B"] = ColumnDimension(sheet, index="B", width=16)
    # "Notes"
    col_dim_holder["C"] = ColumnDimension(sheet, index="C", width=40)
    # "Participants"
    col_dim_holder["D"] = ColumnDimension(sheet, index="D", width=10)
    # "Clients"
    col_dim_holder["E"] = ColumnDimension(sheet, index="E", width=16)
    # "Hours"
    col_dim_holder["F"] = ColumnDimension(sheet, index="F", width=5)
    # "Travel Hours"
    col_dim_holder["G"] = ColumnDimension(sheet, index="G", width=5)
    # "Cumul Hours"
    col_dim_holder["H"] = ColumnDimension(sheet, index="H", width=10)
    sheet.column_dimensions = col_dim_holder

    curr_row = 2   # The header is already in this sheet

    for (date, summary, notes, consultant, client, hours, travel_hours, cumul_hours) in pi_tag_to_consulting_details[pi_tag]:

        curr_col = 1
        sheet.cell(curr_row, curr_col, date).style = DATE_FORMAT; curr_col += 1
        sheet.cell(curr_row, curr_col, summary);           curr_col += 1
        sheet.cell(curr_row, curr_col, notes);             curr_col += 1
        sheet.cell(curr_row, curr_col, consultant);        curr_col += 1
        sheet.cell(curr_row, curr_col, client);            curr_col += 1
        sheet.cell(curr_row, curr_col, hours).style = FLOAT_FORMAT;        curr_col += 1
        sheet.cell(curr_row, curr_col, travel_hours).style = FLOAT_FORMAT; curr_col += 1
        sheet.cell(curr_row, curr_col, cumul_hours).style = FLOAT_FORMAT;  curr_col += 1

        curr_row += 1


# Generates the Lab Users sheet for a BillingNotification workbook with
# username details for a particular PI.  It reads from dicts pi_tag_to_user_details and username_to_user_details.
def generate_lab_users_sheet(sheet, pi_tag):

    # Freeze the first row.
    sheet.freeze_panes = 'A2'

    # Set the column widths
    col_dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    # "Username"
    col_dim_holder["A"] = ColumnDimension(sheet, index="A", width=10)
    # "Full Name"
    col_dim_holder["B"] = ColumnDimension(sheet, index="B", width=20)
    # "Email"
    col_dim_holder["C"] = ColumnDimension(sheet, index="C", width=20)
    # "Date Added"
    col_dim_holder["D"] = ColumnDimension(sheet, index="D", width=10)
    # "Date Removed"
    col_dim_holder["E"] = ColumnDimension(sheet, index="E", width=12)
    sheet.column_dimensions = col_dim_holder

    # Write the user details for active users and moving the inactive users to a separate list.
    past_user_details = []

    # curr_row = 1
    curr_row = 2  # The header is already in this sheet
    for (username, date_added, date_removed, pctage) in pi_tag_to_user_details[pi_tag]:

        curr_col = 1

        # Get the user details for username.
        (email, fullname) = username_to_user_details[username]

        if date_removed == '' or date_removed is None:
            sheet.cell(curr_row, curr_col, username); curr_col += 1
            sheet.cell(curr_row, curr_col, fullname); curr_col += 1
            sheet.cell(curr_row, curr_col, email);    curr_col += 1
            sheet.cell(curr_row, curr_col, date_added).style = DATE_FORMAT; curr_col += 1
            sheet.cell(curr_row, curr_col, "current"); curr_col += 1
            curr_row += 1
        else:
            # Users who have been removed will be listed in a table below the current lab members
            past_user_details.append([username, email, fullname, date_added, date_removed])

    # Write out a subheader for the Previous Lab Members.
    curr_row += 1  # Skip a row before the subheader.
    sheet.cell(curr_row, 1, "Previous Lab Members").style = BOLD_FORMAT
    curr_row += 1
    for (username, email, fullname, date_added, date_removed) in past_user_details:

        curr_col = 1
        sheet.cell(curr_row, curr_col, username); curr_col += 1
        sheet.cell(curr_row, curr_col, fullname); curr_col += 1
        sheet.cell(curr_row, curr_col, email);    curr_col += 1
        sheet.cell(curr_row, curr_col, date_added).style = DATE_FORMAT;   curr_col += 1
        sheet.cell(curr_row, curr_col, date_removed).style = DATE_FORMAT; curr_col += 1

        curr_row += 1


# Generates the Totals sheet for a BillingAggregate workbook, populating the sheet
# from the pi_tag_to_charges dict.
def generate_aggregrate_sheet(sheet):

    # Freeze the first row.
    sheet.freeze_panes = 'A2'

    # Set column widths
    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(sheet)
    dim_holder["A"] = ColumnDimension(sheet, index="A", width=12)
    dim_holder["B"] = ColumnDimension(sheet, index="B", width=12)
    dim_holder["C"] = ColumnDimension(sheet, index="C", width=12)
    dim_holder["D"] = ColumnDimension(sheet, index="D", width=20) # iLab service request name
    dim_holder["E"] = ColumnDimension(sheet, index="E", width=12)
    dim_holder["F"] = ColumnDimension(sheet, index="F", width=12)
    dim_holder["G"] = ColumnDimension(sheet, index="G", width=12)
    dim_holder["H"] = ColumnDimension(sheet, index="H", width=12)
    dim_holder["I"] = ColumnDimension(sheet, index="I", width=12)

    sheet.column_dimensions = dim_holder

    total_fmt = make_format(billing_aggreg_wkbk,
                            {'font_size': 14, 'bold': True})
    charge_fmt = make_format(billing_aggreg_wkbk,
                             {'font_size': 10, 'align': 'right',
                              'num_format': '$#,##0.00'})
    sub_total_charge_fmt = make_format(billing_aggreg_wkbk,
                                       {'font_size': 14, 'align': 'right',
                                        'num_format': '$#,##0.00'})
    grand_charge_fmt = make_format(billing_aggreg_wkbk,
                                   {'font_size': 14, 'align': 'right', 'bold': True,
                                    'num_format': '$#,##0.00'})

    sub_total_storage = 0.0
    sub_total_computing = 0.0
    sub_total_cloud = 0.0
    sub_total_consulting = 0.0
    grand_total_charges = 0.0

    # Compute column numbers for various columns.
    storage_column_num     = BILLING_AGGREG_SHEET_COLUMNS['Totals'].index('Storage') + 1
    computing_column_num   = BILLING_AGGREG_SHEET_COLUMNS['Totals'].index('Computing') + 1
    cloud_column_num       = BILLING_AGGREG_SHEET_COLUMNS['Totals'].index('Cloud') + 1
    consulting_column_num  = BILLING_AGGREG_SHEET_COLUMNS['Totals'].index('Consulting') + 1

    # Sort PI Tags by PI's last name
    pi_tags_sorted = sorted([[pi_tag, pi_tag_to_names_email[pi_tag][1]] for pi_tag in pi_tag_to_charges.keys()],
                            key=lambda a: a[1])

    #curr_row = 1
    curr_row = 2  # Below header
    for pi_tag in [pi_tag_list[0] for pi_tag_list in pi_tags_sorted]:

        (storage, computing, cloud, consulting, total_charges) = pi_tag_to_charges[pi_tag]
        (pi_first_name, pi_last_name, _) = pi_tag_to_names_email[pi_tag]
        (serv_req_id, serv_req_name, serv_req_owner) = pi_tag_to_iLab_info[pi_tag]

        curr_col = 1
        sheet.cell(curr_row, curr_col, pi_first_name);        curr_col += 1
        sheet.cell(curr_row, curr_col, pi_last_name);         curr_col += 1
        sheet.cell(curr_row, curr_col, pi_tag);               curr_col += 1
        sheet.cell(curr_row, curr_col, serv_req_name);        curr_col += 1
        sheet.cell(curr_row, curr_col, storage).style = charge_fmt;  curr_col += 1
        sheet.cell(curr_row, curr_col, computing).style = charge_fmt;curr_col += 1
        sheet.cell(curr_row, curr_col, cloud).style = charge_fmt;        curr_col += 1
        sheet.cell(curr_row, curr_col, consulting).style = charge_fmt;   curr_col += 1

        storage_a1_cell    = rowcol_to_a1_cell(curr_row, storage_column_num)
        computing_a1_cell  = rowcol_to_a1_cell(curr_row, computing_column_num)
        cloud_a1_cell      = rowcol_to_a1_cell(curr_row, cloud_column_num)
        consulting_a1_cell = rowcol_to_a1_cell(curr_row, consulting_column_num)

        sheet.cell(curr_row, curr_col, '=SUM(%s:%s)' % (storage_a1_cell, consulting_a1_cell)).style = charge_fmt
        curr_col += 1

        sub_total_storage += storage
        sub_total_computing += computing
        sub_total_cloud += cloud
        sub_total_consulting += consulting
        grand_total_charges += total_charges

        curr_row += 1

    storage_a1_cell    = rowcol_to_a1_cell(curr_row, storage_column_num)
    computing_a1_cell  = rowcol_to_a1_cell(curr_row, computing_column_num)
    cloud_a1_cell      = rowcol_to_a1_cell(curr_row, cloud_column_num)
    consulting_a1_cell = rowcol_to_a1_cell(curr_row, consulting_column_num)

    sheet.cell(curr_row, 1, "TOTALS").style = total_fmt
    top_storage_a1_cell = rowcol_to_a1_cell(2, storage_column_num)
    bot_storage_a1_cell = rowcol_to_a1_cell(curr_row - 1, storage_column_num)
    sheet.cell(curr_row, storage_column_num , '=SUM(%s:%s)' % (top_storage_a1_cell, bot_storage_a1_cell)).style = sub_total_charge_fmt
    top_computing_a1_cell = rowcol_to_a1_cell(2, computing_column_num)
    bot_computing_a1_cell = rowcol_to_a1_cell(curr_row - 1, computing_column_num)
    sheet.cell(curr_row, computing_column_num, '=SUM(%s:%s)' % (top_computing_a1_cell, bot_computing_a1_cell)).style = sub_total_charge_fmt

    top_cloud_a1_cell = rowcol_to_a1_cell(2, cloud_column_num)
    bot_cloud_a1_cell = rowcol_to_a1_cell(curr_row - 1, cloud_column_num)
    sheet.cell(curr_row, cloud_column_num, '=SUM(%s:%s)' % (top_cloud_a1_cell, bot_cloud_a1_cell)).style = sub_total_charge_fmt

    top_consulting_a1_cell = rowcol_to_a1_cell(2, consulting_column_num)
    bot_consulting_a1_cell = rowcol_to_a1_cell(curr_row - 1, consulting_column_num)
    sheet.cell(curr_row, consulting_column_num, '=SUM(%s:%s)' % (top_consulting_a1_cell, bot_consulting_a1_cell)).style = sub_total_charge_fmt

    sheet.cell(curr_row, consulting_column_num + 1, '=%s+%s+%s+%s' % (storage_a1_cell, computing_a1_cell, cloud_a1_cell, consulting_a1_cell)).style = grand_charge_fmt

#=====
#
# SCRIPT BODY
#
#=====

parser = argparse.ArgumentParser(parents=[argparse_get_parent_parser()])

parser.add_argument("-D","--billing_details_file",
                    default=None,
                    help='The BillingDetails file')
parser.add_argument("-p", "--pi_sheets", action="store_true",
                    default=False,
                    help='Add PI-specific sheets to the BillingAggregate workbook [default = False]')
parser.add_argument("--cpu_time_unit", choices=['cpu-hours', 'cpu-days'],
                    default='cpu-days',
                    help='Choose the CPU time units [default = cpu-days]')

args = parser.parse_args()

#
# Process arguments.
#

# Get year/month-related arguments
(year, month, begin_month_timestamp, end_month_timestamp) = argparse_get_year_month(args)

# Get BillingRoot and BillingConfig arguments
(billing_root, billing_config_file) = argparse_get_billingroot_billingconfig(args, year, month)

###
#
# Read the BillingConfig workbook and build input data structures.
#
###

# billing_config_wkbk = xlrd.open_workbook(billing_config_file)
billing_config_wkbk = openpyxl.load_workbook(billing_config_file)

# Within BillingRoot, create YEAR/MONTH dirs if necessary.
input_subdir = get_subdirectory(billing_root, year, month, "")

# If BillingDetails file given, use that, else look in BillingRoot.
if args.billing_details_file is not None:
    billing_details_file = args.billing_details_file
else:
    billing_details_file = os.path.join(input_subdir, "%s.%s-%02d.xlsx" % (BILLING_DETAILS_PREFIX, year, month))

# Get the absolute path for the billing_details_file.
billing_details_file = os.path.abspath(billing_details_file)

# Build the path to write the Notifications files into
notifs_output_subdir = get_subdirectory(billing_root, year, month, SUBDIR_INVOICES, create_if_nec=True)
# Build the path to write the Aggregate file into
aggregate_output_subdir = get_subdirectory(billing_root, year, month, "")

#
# Output the state of arguments.
#
print("GENERATING NOTIFICATIONS FOR %02d/%d:" % (month, year))
print("  BillingConfigFile: %s" % billing_config_file)
print("  BillingRoot: %s" % billing_root)
print("  BillingDetailsFile: %s" % billing_details_file)
print()

#
# Build data structures.
#
print("Building configuration data structures.")
build_global_data(billing_config_wkbk, begin_month_timestamp, end_month_timestamp)

###
#
# Read the BillingDetails workbook, and create output data structures.
#
###

# Open the BillingDetails workbook.
print("Read in BillingDetails workbook.")
#billing_details_wkbk = xlrd.open_workbook(billing_details_file)
billing_details_wkbk = openpyxl.load_workbook(billing_details_file)

# Read in its Storage sheet and generate output data.
print("Reading storage sheet.")
read_storage_sheet(billing_details_wkbk)

# Read in its Computing sheet and generate output data.
print("Reading computing sheet.")
read_computing_sheet(billing_details_wkbk)

print("Reading cloud sheet.")
read_cloud_sheet(billing_details_wkbk)

# Read in its Consulting sheet and generate output data.
print("Reading consulting sheet.")
read_consulting_sheet(billing_details_wkbk)

###
#
# Write BillingNotification workbooks from output data structures.
#
###

print("Writing notification workbooks:")
for pi_tag in sorted(pi_tag_list):

    print(" %s" % pi_tag)

    # Initialize the BillingNotification spreadsheet for this PI.
    notifs_wkbk_filename = "%s-%s.%s-%02d.xlsx" % (BILLING_NOTIFS_PREFIX, pi_tag, year, month)
    notifs_wkbk_pathname = os.path.join(notifs_output_subdir, notifs_wkbk_filename)

    # billing_notifs_wkbk = xlsxwriter.Workbook(notifs_wkbk_pathname)
    billing_notifs_wkbk = openpyxl.Workbook(write_only=False)
    sheet_name_to_sheet_map = init_billing_notifs_wkbk(billing_notifs_wkbk)

    # Generate the Billing sheet.
    generate_billing_sheet(billing_notifs_wkbk, sheet_name_to_sheet_map['Billing'],
                           pi_tag, begin_month_timestamp, end_month_timestamp)

    # Generate the Rates sheet.
    #generate_rates_sheet(billing_config_wkbk.sheet_by_name('Rates'), sheet_name_to_sheet_map['Rates'])
    generate_rates_sheet(billing_config_wkbk['Rates'], pi_tag, sheet_name_to_sheet_map['Rates'])

    # Generate the Computing Details sheet.
    generate_computing_details_sheet(billing_notifs_wkbk, sheet_name_to_sheet_map['Computing Details'], pi_tag)

    # Generate the Cloud Details sheet.
    generate_cloud_details_sheet(sheet_name_to_sheet_map['Cloud Details'], pi_tag)

    # Generate the Lab Users sheet.
    generate_lab_users_sheet(sheet_name_to_sheet_map['Lab Users'], pi_tag)

    # Generate the Consulting Details
    generate_consulting_details_sheet(sheet_name_to_sheet_map['Consulting Details'], pi_tag)

    billing_notifs_wkbk.save(notifs_wkbk_pathname)

###
#
# Write BillingAggregate workbook from totals in BillingNotifications workbooks.
#
###

print("Writing billing aggregate workbook.")

aggreg_wkbk_filename = "%s.%s-%02d.xlsx" % (BILLING_AGGREGATE_PREFIX, year, month)
aggreg_wkbk_pathname = os.path.join(input_subdir, aggreg_wkbk_filename)

# billing_aggreg_wkbk = xlsxwriter.Workbook(aggreg_wkbk_pathname)
billing_aggreg_wkbk = openpyxl.Workbook()

aggreg_sheet_name_to_sheet = init_billing_aggreg_wkbk(billing_aggreg_wkbk, pi_tag_list)

aggreg_totals_sheet = aggreg_sheet_name_to_sheet['Totals']

# Create the aggregate Totals sheet.
generate_aggregrate_sheet(aggreg_totals_sheet)

if args.pi_sheets:
    # Add the Billing sheets for each PI.
    for pi_tag in sorted(pi_tag_list):

        pi_sheet = aggreg_sheet_name_to_sheet[pi_tag]

        generate_billing_sheet(billing_aggreg_wkbk, pi_sheet,
                               pi_tag, begin_month_timestamp, end_month_timestamp)

billing_aggreg_wkbk.save(aggreg_wkbk_pathname)

###
#
# Output some summary statistics.
#
###
total_jobs_billed = 0
for pi_tag in pi_tag_list:
    total_jobs_billed += len(pi_tag_to_job_details[pi_tag])

print("Total Jobs Billed:", total_jobs_billed)
