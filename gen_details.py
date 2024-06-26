#!/usr/bin/env python3

#===============================================================================
#
# gen_details.py - Generate billing details for month/year.
#
# ARGS:
#   1st: the BillingConfig spreadsheet.
#
# SWITCHES:
#   --accounting_file: Location of accounting file (overrides BillingConfig.xlsx)
#   --billing_root:    Location of BillingRoot directory (overrides BillingConfig.xlsx)
#                      [default if no BillingRoot in BillingConfig.xlsx or switch given: CWD]
#   --year:            Year of snapshot requested. [Default is this year]
#   --month:           Month of snapshot requested. [Default is last month]
#   --no_storage:      Don't run the storage calculations.
#   --no_usage:        Don't run the storage usage calculations (only the quotas).
#   --no_computing:    Don't run the computing calculations.
#   --no_consulting:   Don't run the consulting calculations.
#   --all_jobs_billable: Consider all jobs to be billable. [default=False]
#   --ignore_job_timestamps: Ignore timestamps in job and allow jobs not in month selected [default=False]
#
# INPUT:
#   BillingConfig spreadsheet.
#   SGE Accounting snapshot file (from snapshot_accounting.py).
#     - Expected in BillingRoot/<year>/<month>/SGEAccounting.<year>-<month>.xlsx
#
# OUTPUT:
#   BillingDetails spreadsheet in BillingRoot/<year>/<month>/BillingDetails.<year>-<month>.xlsx
#   Various messages about current processing status to STDOUT.
#
# ASSUMPTIONS:
#   Depends on xlrd and xlsxwriter modules.
#   The input spreadsheet has been certified by check_config.py.
#
# AUTHOR:
#   Keith Bettinger
#
#==============================================================================

#=====
#
# IMPORTS
#
#=====
import argparse
import codecs
import collections
import csv
import datetime
import locale
import os
import os.path
import re
import sys

import openpyxl
import openpyxl.styles

from job_accounting_file import JobAccountingFile


# Simulate an "include billing_common.py".
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
exec(compile(open(os.path.join(SCRIPT_DIR, "billing_common.py"), "rb").read(), os.path.join(SCRIPT_DIR, "billing_common.py"), 'exec'))

#=====
#
# CONSTANTS
#
#=====


#=====
#
# GLOBALS
#
#=====
# In billing_common.py
global SLURMACCOUNTING_PREFIX
global GOOGLE_INVOICE_PREFIX
global BILLING_DETAILS_PREFIX
global CONSULTING_PREFIX
global STORAGE_PREFIX
global ACCOUNT_PREFIXES
global EXCEL_MAX_ROWS
global SUBDIR_RAWDATA

#
# Formats for the output workbook, to be initialized along with the workbook.
#
BOLD_FORMAT = None
DATE_FORMAT = None
INT_FORMAT  = None
FLOAT_FORMAT = None
MONEY_FORMAT = None
PERCENT_FORMAT = None

# Set locale to be US english for converting strings with commas into floats.
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

#=====
#
# FUNCTIONS
#
#=====

# In billing_common.py
global read_config_sheet
global sheet_get_named_column
global sheet_name_to_sheet
global from_timestamp_to_excel_date
global from_excel_date_to_timestamp
global from_timestamp_to_date_string
global from_excel_date_to_date_string
global from_ymd_date_to_timestamp
global from_timestamp_to_datetime
global from_datetime_to_timestamp
global from_excel_date_to_datetime
global from_datetime_to_excel_date
global from_ymd_date_to_datetime
global remove_unicode_chars
global filter_by_dates
global argparse_get_parent_parser
global argparse_get_year_month
global argparse_get_billingroot_billingconfig
global get_subdirectory

# Initialize the output BillingDetails workbook, given as argument.
# It creates all the formats used within the workbook, and saves them
# as the global variables listed at the top of the method.
# It also creates all the sheets within the workbook, and the column
# headers within those sheets.  The method returns a dict of mappings
# from sheet_name to the workbook's Sheet object for that name.
def init_billing_details_wkbk(workbook):

    global BOLD_FORMAT
    global DATE_FORMAT
    global FLOAT_FORMAT
    global INT_FORMAT
    global MONEY_FORMAT
    global PERCENT_FORMAT

    global BILLING_DETAILS_SHEET_COLUMNS  # In billing_common.py

    # Create formats for use within the workbook.
    BOLD_FORMAT = openpyxl.styles.NamedStyle(name="bold")
    BOLD_FORMAT.font = openpyxl.styles.Font(bold=True)
    workbook.add_named_style(BOLD_FORMAT)

    DATE_FORMAT = openpyxl.styles.NamedStyle(name="date")
    DATE_FORMAT.number_format = 'mm/dd/yy'
    workbook.add_named_style(DATE_FORMAT)

    INT_FORMAT = openpyxl.styles.NamedStyle(name="int")
    INT_FORMAT.number_format = '0'
    workbook.add_named_style(INT_FORMAT)

    FLOAT_FORMAT = openpyxl.styles.NamedStyle(name="float")
    FLOAT_FORMAT.number_format = '0.0'
    workbook.add_named_style(FLOAT_FORMAT)

    MONEY_FORMAT = openpyxl.styles.NamedStyle(name="money")
    MONEY_FORMAT.number_format = '$0.00'
    workbook.add_named_style(MONEY_FORMAT)

    PERCENT_FORMAT = openpyxl.styles.NamedStyle(name="percent")
    PERCENT_FORMAT.number_format = '0%'
    workbook.add_named_style(PERCENT_FORMAT)

    sheet_name_to_sheet_map = dict()

    # Control the size of the Workbook when it opens
    view = [openpyxl.workbook.views.BookView(windowWidth=18140, windowHeight=30000)]
    workbook.views = view

    for sheet_name in BILLING_DETAILS_SHEET_COLUMNS:

        sheet = workbook.create_sheet(sheet_name)
        for col in range(0, len(BILLING_DETAILS_SHEET_COLUMNS[sheet_name])):
            sheet.cell(1, col+1, BILLING_DETAILS_SHEET_COLUMNS[sheet_name][col]).style = BOLD_FORMAT

        sheet_name_to_sheet_map[sheet_name] = sheet

    return sheet_name_to_sheet_map


# Given a list of tuples of job details, writes the job details to
# the sheet given.  Possible sheets for use in this method are
# typically the "Computing", "Nonbillable Jobs", and "Failed Jobs" sheets.
def write_job_details(workbook, sheet, sheet_name, job_details):

    # If no job details, write "No Jobs".
    if len(job_details) == 0:
        sheet.write(1, 0, "No jobs")
        print()
        return

    num_jobs = len(job_details)
    job_rows_left = num_jobs

    first_job_row = 0
    last_job_row = first_job_row + min(job_rows_left, EXCEL_MAX_ROWS-1)  # Subtract 1 for header line

    sheet_number = 1

    print(num_jobs)

    while True:

        sheet_row = 0

        # If we have job details, write them to the sheet, below the headers.
        for row in range(first_job_row, last_job_row):

            # Bump rows down below header line.
            sheet_row += 1

            # A little feedback for the people.
            if not args.verbose:
                if sheet_row % 1000 == 0:
                    sys.stdout.write('.')
                    sys.stdout.flush()

            # 'Job Date'
            col = 0
            sheet.cell(sheet_row+1, col+1, from_timestamp_to_excel_date(job_details[row][col])).style = DATE_FORMAT


            # 'Job Timestamp'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col]).style = INT_FORMAT
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Username'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col])
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Job Name'
            col += 1
            sheet.cell(sheet_row+1, col+1, str(job_details[row][col]))
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Account'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col])
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Node'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col])
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Slots'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col]).style = INT_FORMAT
            if args.verbose: print(job_details[row][col], end=' ')

            # 'Wallclock Secs'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col]).style = INT_FORMAT
            if args.verbose: print(job_details[row][col], end=' ')

            # 'JobID'
            col += 1
            sheet.cell(sheet_row+1, col+1, job_details[row][col]).style = INT_FORMAT
            if args.verbose: print(job_details[row][col], end=' ')

            # Extra column if needed: 'Reason' or 'Failed Code'
            if col < len(job_details[row])-1:
                col += 1
                sheet.cell(sheet_row+1, col+1, job_details[row][col])
                if args.verbose: print(job_details[row][col], end=' ')

            if args.verbose: print()

        job_rows_left -= last_job_row - first_job_row

        first_job_row = last_job_row
        last_job_row = first_job_row + min(job_rows_left, EXCEL_MAX_ROWS-1)  # Subtract 1 for header line

        if job_rows_left > 0:
            # Generate new sheet of the form "<sheet name> <sheet number>" (starting with 2).
            sheet_number += 1
            sheet = workbook.create_sheet("%s %d" % (sheet_name, sheet_number))

            # Create same headers on new numbered sheets as on the original sheet name.
            for col in range(0, len(BILLING_DETAILS_SHEET_COLUMNS[sheet_name])):
                sheet.cell(1, col+1, BILLING_DETAILS_SHEET_COLUMNS[sheet_name][col]).style = BOLD_FORMAT
            sys.stdout.write('|')
            sys.stdout.flush()
        else:
            break  # Leave while True loop.

    print()

#
# Returns True/False if string is a valid account.
#
def is_valid_account(acct):

    global account_list
    global pi_tag_list

    # Let's go case insensitive.
    acct = acct.lower()

    # If this is a known account or matches a PI Tag, we are good.
    if acct in account_list or acct in pi_tag_list:
        return True
    # Otherwise, does it match the pattern of <PREFIX>_<PITag>, where PREFIX is in ACCOUNT_PREFIXES?
    else:
        # Split account into underline-separated words.
        ul_words = acct.split('_')

        # The prefix is everything but the last word, which is then expected to be a PI Tag.
        prefix = '_'.join(ul_words[0:-1])
        pi_tag_word = ul_words[-1]

        if prefix in ACCOUNT_PREFIXES and pi_tag_word in pi_tag_list:
            return True
        else:
            print(acct, "is not a valid account (%s %s)." % (prefix, pi_tag_word))
            return False


# Read the Storage Usage file.
# Returns mapping from folders to [timestamp, total, used]
def read_storage_usage_file(storage_usage_file):

    # Mapping from folders to [timestamp, total, used].
    folder_size_dict = collections.OrderedDict()

    usage_fileobj = open(storage_usage_file)

    usage_csvdict = csv.DictReader(usage_fileobj)

    for row in usage_csvdict:
        folder_size_dict[row['Folder']] = \
            (float(row['Timestamp']), float(row['Size']), float(row['Used']),
             int(row['Inodes Quota']), int(row['Inodes Used']))

    return folder_size_dict


# Write storage usage data into the Storage sheet of the BillingDetails file.
# Takes in mapping from folders to [timestamp, total, used].
def write_storage_usage_data(folder_size_dict, storage_sheet):

    # Write space-used mapping into details workbook.
    row = 0
    for folder in sorted(list(folder_size_dict.keys())):

        [timestamp, total, used, inodes_quota, inodes_used] = folder_size_dict[folder]
        sheet_row = row + 1

        # 'Date Measured'
        col = 0
        storage_sheet.cell(sheet_row+1, col+1, from_timestamp_to_excel_date(timestamp)).style = DATE_FORMAT

        # 'Timestamp'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, timestamp)
        if args.verbose: print(timestamp, end=' ')

        # 'Folder'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, folder)
        if args.verbose: print(folder, end=' ')

        # 'Size'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, total).style = FLOAT_FORMAT
        if args.verbose: print(total, end=' ')

        # 'Used'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, used).style = FLOAT_FORMAT
        if args.verbose: print(used, end=' ')

        # 'Inodes Quota'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, inodes_quota).style = INT_FORMAT
        if args.verbose: print(inodes_quota, end=' ')

        # 'Inodes Used'
        col += 1
        storage_sheet.cell(sheet_row+1, col+1, inodes_used).style = INT_FORMAT
        if args.verbose: print(inodes_used)

        # Next row, please.
        row += 1


# Generates the job details stored in the "Computing", "Nonbillable Jobs", and "Failed Jobs" sheets.
def compute_computing_charges(config_wkbk, begin_timestamp, end_timestamp, accounting_file,
                              computing_sheet, nonbillable_job_sheet, failed_job_sheet):

    # In billing_common.py
    global ACCOUNTING_FIELDS
    global ACCOUNTING_FAILED_CODES
    global BILLABLE_HOSTNAME_PREFIXES
    global NONBILLABLE_HOSTNAME_PREFIXES
    global IGNORED_ACCOUNTS
    global account_list
    global pi_tag_list

    print("Computing computing charges...")

    # Read in the Usernames from the Users sheet.
    users_sheet = config_wkbk['Users']
    users_list = sheet_get_named_column(users_sheet, "Username")
    #  NOTE: This column may have some duplicates in it.
    #        Need to make a set out of the result.
    users_list = set(users_list)

    print("  Reading accounting file %s" % (os.path.abspath(accounting_file)))

    #
    # Open the current accounting file for input.
    #
    accounting_fp = JobAccountingFile(accounting_file)

    #
    # Read all the lines of the current accounting file.
    #  Output to the details spreadsheet those jobs
    #  which have "end_times" in the given month,
    #  and "owner"s in the list of users.
    #
    not_in_users_list = set()
    not_in_account_list = collections.defaultdict(set)
    both_proj_and_acct_list = collections.defaultdict(set)

    failed_job_details           = []  # Jobs which failed.
    billable_job_details         = []  # Jobs that are on hosts we can bill for.
    nonbillable_node_job_details = []  # Jobs not on hosts we can bill for.
    unknown_node_job_details     = []  # Jobs on unknown nodes.
    both_billable_and_non_node_job_details = []  # Jobs which have both billable and nonbillable nodes.
    unknown_user_job_details     = []  # Jobs from users we don't know.

    unknown_job_nodes            = set()  # Set of nodes we don't know.

    jobids_with_unknown_billable_nodes = set()  # Set of job IDs for jobs which have nodes that can't be identified as billable.
    jobids_with_billable_and_non_nodes = set()  # Set of job IDs for jobs which have both billable and nonbillable nodes.

    for accounting_record in accounting_fp:

        # If the job failed, the submission_time is the job date.
        # Else, the end_time is the job date.
        failed_code = accounting_record.failed_code
        job_failed = failed_code in ACCOUNTING_FAILED_CODES
        if job_failed:
            job_date = accounting_record.submission_time  # The only valid date in the record.
        else:
            job_date = accounting_record.end_time

        # Create a list of job details for this job.
        job_details = list()
        job_details.append(job_date)
        job_details.append(job_date)  # Two columns used for the date: one date formatted, one timestamp.
        job_details.append(accounting_record.owner)
        job_details.append(accounting_record.job_name)

        #
        # Look for accounts in both account and project fields.
        # If values occur in both, use the project field and record the discrepancy.
        #
        job_account = remove_unicode_chars(accounting_record.account)
        if job_account == 'sge' or job_account == '':   # Edit out the default account 'sge'.
            job_account = None

        job_project = remove_unicode_chars(accounting_record.project)
        if job_project == 'NONE' or job_project == '':  # Edit out the placeholder project 'NONE'.
            job_project = None

        #
        # Add account (project/account) info to job_details.
        #
        # If project is set and not in the ignored account list:
        if job_project is not None and job_project not in IGNORED_ACCOUNTS:

            # Find out if the project name is a known one.
            job_project_is_valid_account = is_valid_account(job_project)

            if not job_project_is_valid_account:
                # If this project/account is unknown, save details for later output.
                not_in_account_list[accounting_record.owner].add(job_project)
        else:
            job_project_is_valid_account = False
            job_project = None  # we could be ignoring a given account

        # If account is set and not in the ignored account list:
        if job_account is not None and job_account not in IGNORED_ACCOUNTS:

            # Find out if the account name is a known one.
            job_account_is_valid_account = is_valid_account(job_account)

            if not job_account_is_valid_account:
                # If this account is unknown, save details for later output.
                not_in_account_list[accounting_record.owner].add(job_account)
        else:
            job_account_is_valid_account = False
            job_account = None  # we could be ignoring a given account

        #
        # Decide which of project and account will be used for account.
        #

        # If project is valid, choose project for account.
        if job_project_is_valid_account:

            # If there's both a project and an account, choose the project and save details for later output.
            job_account = job_project
            if job_account is not None:
                both_proj_and_acct_list[accounting_record.owner].add((job_project,job_account))

        # Else if project is present and account is not valid, choose project for account.
        # (Non-valid project trumps non-valid account).
        elif job_project is not None and not job_account_is_valid_account:

            # If there's both a project and an account, choose the project and save details for later output.
            job_account = job_project
            if job_account is not None:
                both_proj_and_acct_list[accounting_record.owner].add((job_project,job_account))

        # Else if account is present, choose account for account.
        # (either account is valid and the project is non-valid, or there is no project).
        elif job_account is not None:
            job_account = job_account

            # If there's both an account and a project, save the details for later output.
            if job_project is not None:
                both_proj_and_acct_list[accounting_record.owner].add((job_project,job_account))

        # else No project and No account = No account.
        else:
            job_account = None

        # Add the computed account to the job_details, if any.
        if job_account is not None:
            job_details.append(job_account)
        else:
            job_details.append('')

        # Support for Slurm: 'hostname' is now a comma-separated node list.
        node_list = accounting_record.node_list
        # Edit hostname to remove trailing ".local".
        node_list = node_list.replace(".local","")
        job_details.append(node_list)

        job_details.append(accounting_record.cpus)

        wallclock = accounting_record.wallclock  # run time in seconds
        job_details.append(wallclock)

        job_details.append(accounting_record.job_id)

        # If the end date of this job was within the month or we aren't reading job timestamps,
        #  examine it.
        if args.ignore_job_timestamps or begin_timestamp <= job_date < end_timestamp:

            job_nodes_are_billable    = list()
            job_nodes_are_nonbillable = list()

            # Is the job's node billable?
            if not args.all_jobs_billable:

                # Need to convert commas to semicolons in lists marked by [ ]'s
                match_bracket_lists = re.findall('(\[.*?\]+)',node_list)

                for substr in match_bracket_lists:
                    new_substr = substr.replace(',', ';')
                    node_list = node_list.replace(substr,new_substr)

                # Now, with the commas only separating the node, we can split the node list by commas.
                list_of_nodes = node_list.split(',')

                for node_name in list_of_nodes:

                    # Put the commas back for an individual node.
                    node_name = node_name.replace(';',',')

                    # Job is billable if it ran on a host starting with one of the BILLABLE_HOSTNAME_PREFIXES.
                    billable    = any([node_name.startswith(p) for p in BILLABLE_HOSTNAME_PREFIXES])
                    # Job is not billable if it ran on a host starting with one of the NONBILLABLE_HOSTNAME_PREFIXES.
                    nonbillable = any([node_name.startswith(p) for p in NONBILLABLE_HOSTNAME_PREFIXES])

                    # Screen for cases where a node is either billable and nonbillable or neither.
                    if billable and nonbillable:
                        print("*** Error: Node %s of Job %s is both billable and non-billable" % (node_name, accounting_record.job_id), file=sys.stderr)
                        jobids_with_billable_and_non_nodes.add(accounting_record.job_id)
                    elif not (billable or nonbillable):
                        print("*** Error: Node %s of Job %s is neither billable nor non-billable" % (node_name, accounting_record.job_id), file=sys.stderr)
                        jobids_with_unknown_billable_nodes.add(accounting_record.job_id)

                    job_nodes_are_billable.append(billable)
                    job_nodes_are_nonbillable.append(nonbillable)

                job_is_billable    = any(job_nodes_are_billable)
                job_is_nonbillable = any(job_nodes_are_nonbillable)
            else:
                job_is_billable    = True
                job_is_nonbillable = False

            job_is_both_billable_and_non = job_is_billable and job_is_nonbillable
            job_is_unknown_billable = not (job_is_billable or job_is_nonbillable)

            # Do we know this job's user?
            job_user_is_known = accounting_record.owner in users_list
            # If not, save the username in an unknown-user list.
            if not job_user_is_known:
                # Save unknown user and job details in unknown user lists.
                not_in_users_list.add(accounting_record.owner)

            # If we know the user or the job has a account...
            if job_user_is_known or job_account is not None:

                # If job failed, save in Failed job list.
                if job_failed:
                    failed_job_details.append(job_details + [failed_code])
                else:
                    # If hostname doesn't have a billable prefix, save in an nonbillable list.
                    if job_is_both_billable_and_non:
                        both_billable_and_non_node_job_details.append(job_details + ['Both Billable and Non Nodes'])
                    elif job_is_unknown_billable:
                        unknown_node_job_details.append(job_details + ['Unknown Node'])
                        unknown_job_nodes.add(node_list)
                    elif job_is_billable:
                        billable_job_details.append(job_details)
                    elif job_is_nonbillable:
                        nonbillable_node_job_details.append(job_details + ['Nonbillable Node'])
                    else:
                        print("  *** Pathological state for job %s billingness. *** " % (accounting_record.job_id))
            else:
                # Save the job details in an unknown-user list.
                unknown_user_job_details.append(job_details + ['Unknown User'])

        else:
            if job_date != 0 and job_date is not None:
                dates_tuple = (from_timestamp_to_date_string(job_date),
                               from_timestamp_to_date_string(begin_timestamp),
                               from_timestamp_to_date_string(end_timestamp))
                print("Job date %s is not between %s and %s" % dates_tuple)
            else:
                print("Job date is zero/None.")

    #
    # ERROR FLAGGING:
    #

    # Print out list of users who had jobs but were not in any lab list.
    if len(not_in_users_list) > 0:
        print("  *** Job submitters not in users list:", end=' ')
        for user in not_in_users_list:
            print(user, end=' ')
        print()
    # Print out list of unknown accounts.
    if len(list(not_in_account_list.keys())) > 0:
        print("  *** Jobs with unknown accounts:")
        for user in sorted(not_in_account_list.keys()):
            print('   ', user)
            for job_account in sorted(not_in_account_list[user]):
                print('     ', job_account)
    # Print out list of jobs with both project and account accounts.
    if len(list(both_proj_and_acct_list.keys())) > 0:
        print("  *** Jobs with both project and account accounts:")
        for user in sorted(both_proj_and_acct_list.keys()):
            print('   ', user)
            for (proj, acct) in both_proj_and_acct_list[user]:
                print('     Project:', proj, 'Account:', acct)
    # Print out how many jobs were run on unknown nodes.
    if len(unknown_job_nodes) > 0:
        print("  *** Unknown Nodes with jobs:")
        for node in sorted(unknown_job_nodes):
            print('   ', node)
    # Print out job IDs of jobs which have both billable and nonbillable nodes.
    if len(jobids_with_billable_and_non_nodes) > 0:
        print("  *** Job IDs with both billable and nonbillable nodes:")
        for jobid in jobids_with_billable_and_non_nodes:
            print('   ', jobid)
    # Print out job IDs of jobs which have nodes which are neither billable nor nonbillable.
    if len(jobids_with_unknown_billable_nodes) > 0:
        print("  *** Job IDs with nodes which are neither billable nor nonbillable:")
        for jobid in jobids_with_unknown_billable_nodes:
            print('   ', jobid)

    # Output the accounting details to the BillingDetails worksheet.
    print("  Outputting accounting details")

    # Output jobs to sheet for billable jobs.
    if len(billable_job_details) > 0:
        print("    Billable Jobs:    ", end=' ')
        write_job_details(billing_details_wkbk, computing_sheet, "Computing", billable_job_details)

    # Output nonbillable jobs to sheet for nonbillable jobs.
    if len(nonbillable_node_job_details) > 0:
        print("    Nonbillable Jobs: ", end=' ')
        all_nonbillable_job_details = \
            nonbillable_node_job_details + unknown_user_job_details + unknown_node_job_details + both_billable_and_non_node_job_details
        write_job_details(billing_details_wkbk, nonbillable_job_sheet, "Nonbillable Jobs", all_nonbillable_job_details)

    # Output jobs to sheet for failed jobs.
    if len(failed_job_details) > 0:
        print("    Failed Jobs:      ", end=' ')
        write_job_details(billing_details_wkbk, failed_job_sheet, "Failed Jobs", failed_job_details)

    print("Computing charges computed.")


# Generates the "Consulting" sheet.
def compute_consulting_charges(config_wkbk, begin_timestamp, end_timestamp, consulting_timesheet, consulting_sheet):

    print("Computing consulting charges...")

    begin_datetime = datetime.datetime.utcfromtimestamp(begin_timestamp)
    end_datetime   = datetime.datetime.utcfromtimestamp(end_timestamp)

    ###
    # Read the config workbook to get a list of active PIs
    ###
    pis_sheet = config_wkbk["PIs"]

    pis_list    = sheet_get_named_column(pis_sheet, "PI Tag")
    dates_added = sheet_get_named_column(pis_sheet, "Date Added")
    dates_remvd = sheet_get_named_column(pis_sheet, "Date Removed")

    # Note: Previous versions had a bug here, passing begin_timestamp and end_timestamp directly to filter_by_dates()
    active_pis_list = filter_by_dates(pis_list, list(zip(dates_added, dates_remvd)), begin_datetime, end_datetime)

    ###
    # Read the Consulting Timesheet.
    ###

    #consulting_workbook = xlrd.open_workbook(consulting_timesheet)
    consulting_workbook = openpyxl.load_workbook(consulting_timesheet, data_only=True) # read_only=True)

    #hours_sheet = consulting_workbook.sheet_by_name("Hours")
    hours_sheet = consulting_workbook["Hours"]

    dates   = sheet_get_named_column(hours_sheet, "Date")
    pi_tags = sheet_get_named_column(hours_sheet, "PI Tag")
    hours   = sheet_get_named_column(hours_sheet, "Hours")
    travel_hours = sheet_get_named_column(hours_sheet, "Travel Hours")
    participants = sheet_get_named_column(hours_sheet, "Participants")
    clients = sheet_get_named_column(hours_sheet, "Clients")
    summaries = sheet_get_named_column(hours_sheet, "Summary")
    notes   = sheet_get_named_column(hours_sheet, "Notes")
    cumul_hours = sheet_get_named_column(hours_sheet, "Cumul Hours")

    # Mar 2018: new column denoting that these entries should be ignored
    # (this entries are paid for by FTE% and not hourly).
    sdrc_members = sheet_get_named_column(hours_sheet, "SDRC ?")
    # If there is no "SDRC?" column (backward compatibility),
    # just make a list of empty strings to zip with the columns above.
    if sdrc_members is None:
        sdrc_members = [""] * len(dates)

    # Convert empty travel hours to zeros.
    travel_hours = [0 if h=='' else h for h in travel_hours]

    row = 1
    for (date, pi_tag, hours_spent, travel_hrs, participant, client, summary, note, cumul_hours_spent, sdrc_member) in \
            zip(dates, pi_tags, hours, travel_hours, participants, clients, summaries, notes, cumul_hours, sdrc_members):

        # If date is blank, we are done.
        if date == "" or date is None:
            break

        # Ignore this entry if there is an X in the "SDRC ?" column.
        if sdrc_member == "X":
            continue

        # Confirm date is within this month.
        try:
            # date_timestamp = from_excel_date_to_timestamp(date)
            date_timestamp = from_datetime_to_timestamp(date)
        except:
            print("Date Error:", date, "Summary:", summary, "Hours:", hours_spent, "Cumul:", cumul_hours_spent, file=sys.stderr)
            continue

        # If date is before beginning of the month or after the end of the month, skip this entry.
        if begin_timestamp > date_timestamp or date_timestamp >= end_timestamp: continue

        # Confirm that pi_tag is in the list of active PIs for this month.
        if pi_tag not in active_pis_list:
            print("  PI %s not in active PI list...skipping" % pi_tag)

        # Copy the entry into the output consulting sheet.
        col = 0
        #consulting_sheet.write(row, col, date, DATE_FORMAT)
        consulting_sheet.cell(row+1, col+1, date).style = DATE_FORMAT
        col += 1
        #consulting_sheet.write(row, col, pi_tag)
        consulting_sheet.cell(row+1, col+1, pi_tag)
        col += 1
        #consulting_sheet.write(row, col, float(hours_spent), FLOAT_FORMAT)
        consulting_sheet.cell(row+1, col+1, float(hours_spent)).style = FLOAT_FORMAT
        col += 1
        #consulting_sheet.write(row, col, float(travel_hrs), FLOAT_FORMAT)
        if travel_hrs is not None:
            consulting_sheet.cell(row+1, col+1, float(travel_hrs)).style = FLOAT_FORMAT
        col += 1
        #consulting_sheet.write(row, col, participant)
        consulting_sheet.cell(row+1, col+1, participant)
        col += 1
        #consulting_sheet.write(row, col, client)
        consulting_sheet.cell(row+1, col+1, client)
        col += 1
        #consulting_sheet.write(row, col, summary)
        consulting_sheet.cell(row+1, col+1, summary)
        col += 1
        #consulting_sheet.write(row, col, note)
        consulting_sheet.cell(row+1, col+1, note)
        col += 1
        try:
            consulting_sheet.cell(row+1, col+1, float(cumul_hours_spent)).style = FLOAT_FORMAT
        except ValueError:
            if cumul_hours_spent == '#NAME?':
                print("Entry dated {0} for '{1}' for PI {2} has uncalculated cumul_hours".format(from_timestamp_to_date_string(from_datetime_to_timestamp(date)), summary, pi_tag))
                sys.exit(-1)

        col += 1

        row += 1


def write_cloud_details_V1(cloud_sheet, row_dict, output_row):

    output_col = 1 # 0
    total_amount = 0.0

    # Write Google data into Cloud sheet.

    # Output 'Platform' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Product'])
    output_col += 1

    # Output 'Account' Field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Order'])
    output_col += 1

    # Output 'Project' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Source'])
    output_col += 1

    # Output 'Description' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Description'])
    output_col += 1

    # Output 'Dates' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Interval'])
    output_col += 1

    # Parse quantity.
    if len(row_dict['Quantity']) > 0:
        quantity = locale.atof(row_dict['Quantity'])
    else:
        quantity = ''

    # Output 'Quantity' field.
    cloud_sheet.cell(output_row+1, output_col+1, quantity).style = FLOAT_FORMAT
    output_col += 1

    # Output 'Unit of Measure' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['UOM'])
    output_col += 1

    # Parse charge.
    amount = locale.atof(row_dict['Amount'])
    # Accumulate total charges.
    total_amount += amount

    # Output 'Charge' field.
    cloud_sheet.cell(output_row+1, output_col+1, amount, MONEY_FORMAT)
    output_col += 1

    return total_amount


def write_cloud_details_V2(cloud_sheet, row_dict, output_row):

    output_col = 0
    total_amount = 0.0

    # Write Google data into Cloud sheet.

    # Output 'Platform' field.
    cloud_sheet.cell(output_row+1, output_col+1, "Google Cloud Platform")
    output_col += 1

    # Output 'Account' field. (subaccount)
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Account ID'])
    output_col += 1

    # Output 'Project' field.  (Project Name + Project ID)
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Source'])
    output_col += 1

    # Output 'Description' field. (SKU description of the charge)
    sku_description = "%s %s" % (row_dict['Product'], row_dict['Resource Type'])
    cloud_sheet.cell(output_row+1, output_col+1, sku_description)
    output_col += 1

    # Output 'Dates' field.
    date_range = "%s-%s" % (row_dict['Start Date'], row_dict['End Date'])
    cloud_sheet.cell(output_row+1, output_col+1, date_range)
    output_col += 1

    # Parse quantity.
    quantity_str = row_dict['Quantity'].strip()
    if len(quantity_str) > 0:
        quantity = locale.atof(quantity_str)
    else:
        quantity = ''

    # Output 'Quantity' field.
    cloud_sheet.cell(output_row+1, output_col+1, quantity).style = FLOAT_FORMAT
    output_col += 1

    # Output 'Unit of Measure' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Unit'])
    output_col += 1

    # Parse charge.
    amount = locale.atof(row_dict['Amount'])
    # Accumulate total charges.
    total_amount += amount

    # Output 'Charge' field.
    cloud_sheet.cell(output_row+1, output_col+1, amount).style = MONEY_FORMAT
    output_col += 1

    return total_amount


def write_cloud_details_V3(cloud_sheet, row_dict, output_row):

    # If Cost type is not Usage, then ignore line.
    if row_dict['Cost type'] != "Usage":
        return 0.0

    output_col = 0
    total_amount = 0.0

    # Write Google data into Cloud sheet.

    # Output 'Platform' field.
    cloud_sheet.cell(output_row+1, output_col+1, "Google Cloud Platform")
    output_col += 1

    service = row_dict['Service description']

    project_id = row_dict['Project ID']
    account    = row_dict['Billing account ID']

    # Output 'Account' field. (subacccount)
    cloud_sheet.cell(output_row+1, output_col+1, account)
    output_col += 1

    # Output 'Project' field.  (Project Name + Project ID)
    cloud_sheet.cell(output_row+1, output_col+1, project_id)
    output_col += 1

    # Output 'Description' field. (SKU description of the charge)
    sku_description = "%s %s" % (service, row_dict['SKU description'])
    cloud_sheet.cell(output_row+1, output_col+1, sku_description)
    output_col += 1

    # Output 'Dates' field.
    date_range = "%s-%s" % (row_dict['Usage start date'], row_dict['Usage end date'])
    cloud_sheet.cell(output_row+1, output_col+1, date_range)
    output_col += 1

    # Parse quantity.
    quantity_str = row_dict['Usage amount'].strip()
    if len(quantity_str) > 0:
        quantity = locale.atof(quantity_str)
    else:
        quantity = ''

    # Output 'Quantity' field.
    cloud_sheet.cell(output_row+1, output_col+1, quantity).style = FLOAT_FORMAT
    output_col += 1

    # Output 'Unit of Measure' field.
    cloud_sheet.cell(output_row+1, output_col+1, row_dict['Usage unit'])
    output_col += 1

    # Parse charge.
    amount = locale.atof(row_dict['Cost ($)'])
    # Accumulate total charges.
    total_amount += amount

    # Output 'Charge' field.
    cloud_sheet.cell(output_row+1, output_col+1, amount).style = MONEY_FORMAT
    output_col += 1

    return total_amount


# Generates the "Cloud" sheet.
def compute_cloud_charges(config_wkbk, google_invoice_csv, cloud_sheet):

    print("Computing cloud charges...")

    google_invoice_version = "V3"  # Hardcoded to only work with latest version

    ###
    # Read the Google Invoice CSV File
    ###

    # Google Invoice CSV files are Unicode with BOM.
    google_invoice_csv_file_obj = codecs.open(google_invoice_csv, encoding='utf-8-sig')

    # Consume summary table at top of file
    if google_invoice_version == "V3":
        # Read lines until line that starts with "Total amount"
        for line in google_invoice_csv_file_obj:
            first_field = line.split(',')[0]
            if first_field == "Total amount due":
                break

    # Accumulate the total amount of charges while processing each line,
    #  to compare with total amount in header in google_invoice_amount_due above.
    google_invoice_total_amount = 0.0

    output_row = 1  # Keeps track of output row in Cloud sheet; starts at 1, below header.

    #   Create CSVReader from subtable
    google_invoice_subtable_csvreader = csv.DictReader(google_invoice_csv_file_obj)

    #   Foreach row in CSVReader
    for row_dict in google_invoice_subtable_csvreader:

        if google_invoice_version == 'V1':
            row_amount = write_cloud_details_V1(cloud_sheet, row_dict, output_row)
            if args.verbose: print(".", end=' ')
        elif google_invoice_version == 'V2':
            row_amount = write_cloud_details_V2(cloud_sheet, row_dict, output_row)
            if args.verbose: print(".", end=' ')
        elif google_invoice_version == 'V3':
            row_amount = write_cloud_details_V3(cloud_sheet, row_dict, output_row)
            if args.verbose: print(".", end=' ')

        # Add up the row charges to compare to total invoice amount.
        google_invoice_total_amount += row_amount

        # Move to next row.
        output_row += 1

    if args.verbose:
        print()
        print("  Google Cloud Total Amount: %5.2f" % (google_invoice_total_amount))


#=====
#
# SCRIPT BODY
#
#=====

parser = argparse.ArgumentParser(parents=[argparse_get_parent_parser()])

parser.add_argument("-a", "--accounting_file",
                    default=None,
                    help='The compute accounting file to read [default = None]')
parser.add_argument("-g", "--google_invoice_csv",
                    default=None,
                    help="The Google Invoice CSV file")
parser.add_argument("-c", "--consulting_timesheet",
                    default=None,
                    help="The consulting timesheet XSLX file.")
parser.add_argument("-s", "--storage_usage_csv",
                    default=None,
                    help="The storage usage CSV file.")
parser.add_argument("--no_storage", action="store_true",
                    default=False,
                    help="Don't run storage calculations [default = false]")
parser.add_argument("--no_usage", action="store_true",
                    default=False,
                    help="Don't run storage usage calculations [default = false]")
parser.add_argument("--no_computing", action="store_true",
                    default=False,
                    help="Don't run computing calculations [default = false]")
parser.add_argument("--no_consulting", action="store_true",
                    default=False,
                    help="Don't run consulting calculations [default = false]")
parser.add_argument("--no_cloud", action="store_true",
                    default=False,
                    help="Don't run cloud calculations [default = false]")
parser.add_argument("--all_jobs_billable", action="store_true",
                    default=False,
                    help="Consider all jobs to be billable [default = false]")
parser.add_argument("-i", "--ignore_job_timestamps", action="store_true",
                    default=False,
                    help="Ignore timestamps in job (and allow jobs not in month selected) [default = false]")

args = parser.parse_args()

#
# Process arguments.
#

# Get year/month-related arguments
(year, month, begin_month_timestamp, end_month_timestamp) = argparse_get_year_month(args)

# Get BillingRoot and BillingConfig arguments
(billing_root, billing_config_file) = argparse_get_billingroot_billingconfig(args, year, month)

# Open the BillingConfig workbook
billing_config_wkbk = openpyxl.load_workbook(billing_config_file)  # , read_only=True)

# Get the path to the input files to be read.
input_subdir = get_subdirectory(billing_root, year, month, SUBDIR_RAWDATA)
# Get the path for the BillingDetails output file
output_subdir = get_subdirectory(billing_root, year, month, "")

# Use switch arg for accounting_file if present, else use file in BillingRoot.
if args.accounting_file is not None:
    accounting_file = args.accounting_file
else:
    accounting_filename = "%s.%d-%02d.txt" % (SLURMACCOUNTING_PREFIX, year, month)
    accounting_file = os.path.join(input_subdir, accounting_filename)

# Get absolute path for accounting_file.
accounting_file = os.path.abspath(accounting_file)

# Use switch arg for google_invoice_file if present, else use file in BillingRoot.
if args.google_invoice_csv is not None:
    google_invoice_csv = args.google_invoice_csv
else:
    google_invoice_filename = "%s.%d-%02d.csv" % (GOOGLE_INVOICE_PREFIX, year, month)
    google_invoice_csv = os.path.join(input_subdir, google_invoice_filename)

# Get absolute path for google_invoice_csv file.
google_invoice_csv = os.path.abspath(google_invoice_csv)

# Confirm that Google Invoice CSV file exists.
if not os.path.exists(google_invoice_csv):
    google_invoice_csv = None

# Use switch arg to read in storage usage file if given.
#  If not given, generate storage data by analyzing folders.
if args.storage_usage_csv is not None:
    storage_usage_file = args.storage_usage_csv
else:
    storage_usage_filename = "%s.%d-%02d.csv" % (STORAGE_PREFIX, year, month)
    storage_usage_file = os.path.join(input_subdir, storage_usage_filename)

# Get absolute path for storage_usage_file.
storage_usage_file = os.path.abspath(storage_usage_file)

# Confirm that the storage usage file exists.
if not os.path.exists(storage_usage_file):
    storage_usage_file = None

# Use switch arg for consulting_timesheet if present, else use file in BillingRoot.
if args.consulting_timesheet is not None:
    consulting_timesheet = args.consulting_timesheet
else:
    consulting_filename = "%s.%d-%02d.xlsx" % (CONSULTING_PREFIX, year, month)
    consulting_timesheet = os.path.join(input_subdir, consulting_filename)

# Get absolute path for consulting_timesheet file.
consulting_timesheet = os.path.abspath(consulting_timesheet)

# Confirm that the consulting_timesheet file exists.
if not os.path.exists(consulting_timesheet):
    consulting_timesheet = None

# Initialize the BillingDetails output spreadsheet.
details_wkbk_filename = "%s.%s-%02d.xlsx" % (BILLING_DETAILS_PREFIX, year, month)
details_wkbk_pathname = os.path.join(output_subdir, details_wkbk_filename)

#billing_details_wkbk = xlsxwriter.Workbook(details_wkbk_pathname)
billing_details_wkbk = openpyxl.Workbook()
billing_details_wkbk.remove(billing_details_wkbk['Sheet'])

# Create all the sheets in the output spreadsheet.
sheet_name_to_sheet_map = init_billing_details_wkbk(billing_details_wkbk)

#
# Output the state of arguments.
#
print("GETTING DETAILS FOR %02d/%d:" % (month, year))
print("  BillingConfigFile: %s" % billing_config_file)
print("  BillingRoot: %s" % billing_root)

if args.no_storage:
    print("  Skipping storage calculations")
    skip_storage = True
elif storage_usage_file is not None:
    print("  Storage usage file: %s" % storage_usage_file)
    skip_storage = False
else:
    print("  No storage usage file given...skipping storage calculations")
    skip_storage = True

if args.no_computing:
    print("  Skipping computing calculations")
    skip_computing = True
elif accounting_file is not None:
    print("  SGEAccountingFile: %s" % accounting_file)
    skip_computing = False
else:
    print("  No accounting file given...skipping computing calculations")
    skip_computing = True
if args.all_jobs_billable:
    print("  All jobs billable.")

if args.no_consulting:
    print("  Skipping consulting calculations")
    skip_consulting = True
elif consulting_timesheet is not None:
    print("  Consulting Timesheet: %s" % consulting_timesheet)
    skip_consulting = False
else:
    print("  No consulting timesheet given...skipping consulting calculations")
    skip_consulting = True

if args.no_cloud:
    print("  Skipping cloud calculations")
    skip_cloud = True
elif google_invoice_csv is not None:
    print("  Google Invoice File: %s" % google_invoice_csv)
    skip_cloud = False
else:
    print("  No Google Invoice file given...skipping cloud calculations")
    skip_cloud = True

print()
print("  BillingDetailsFile to be output: %s" % details_wkbk_pathname)
print()

# Read in the PI Tag list from the PIs sheet.
#pis_sheet = billing_config_wkbk.sheet_by_name('PIs')
pis_sheet = billing_config_wkbk['PIs']
pi_tag_list = sheet_get_named_column(pis_sheet, 'PI Tag')

# Read in the accounts from the accounts sheet.
#accounts_sheet = billing_config_wkbk.sheet_by_name('Accounts')
accounts_sheet = billing_config_wkbk['Accounts']
account_list = sheet_get_named_column(accounts_sheet, 'Account')

#
# Compute storage charges.
#
if not skip_storage:
    folder_usage_dict = read_storage_usage_file(storage_usage_file)

    # Write storage usage data to BillingDetails file.
    write_storage_usage_data(folder_usage_dict, sheet_name_to_sheet_map['Storage'])

#
# Compute computing charges.
#
if not skip_computing:
    compute_computing_charges(billing_config_wkbk, begin_month_timestamp, end_month_timestamp, accounting_file,
                              sheet_name_to_sheet_map['Computing'],
                              sheet_name_to_sheet_map['Nonbillable Jobs'], sheet_name_to_sheet_map['Failed Jobs'])

#
# Compute consulting charges.
#
if not skip_consulting:
     compute_consulting_charges(billing_config_wkbk, begin_month_timestamp, end_month_timestamp, consulting_timesheet,
                                sheet_name_to_sheet_map['Consulting'])

#
# Compute cloud charges.
#
if not skip_cloud:
    compute_cloud_charges(billing_details_wkbk, google_invoice_csv, sheet_name_to_sheet_map['Cloud'])

#
# Close the output workbook and write the .xlsx file.
#
print("Closing BillingDetails workbook....")
billing_details_wkbk.save(details_wkbk_pathname)
print("BILLING DETAILS WORKBOOK COMPLETE")