#!/usr/bin/env python3

#===============================================================================
#
# gen_ilab_upload.py - Generate billing data for upload into iLab.
#
# ARGS:
#   1st: the BillingConfig spreadsheet.
#
# SWITCHES:
#   --billing_details_file: Location of the BillingDetails.xlsx file (default=look in BillingRoot/<year>/<month>)
#   --billing_root:    Location of BillingRoot directory (overrides BillingConfig.xlsx)
#                      [default if no BillingRoot in BillingConfig.xlsx or switch given: CWD]
#   --year:            Year of snapshot requested. [Default is this year]
#   --month:           Month of snapshot requested. [Default is last month]
#
# OUTPUT:
#   CSV file with billing data suitable for uploading into iLab.
#   Various messages about current processing status to STDOUT.
#
# ASSUMPTIONS:
#
# AUTHOR:
#   Keith Bettinger
#
#==============================================================================

#=====
#
# IMPORTS
#
#=====
import argparse
import codecs
from collections import defaultdict
import csv
import locale  # for converting strings with commas into floats
import os
import re
import sys

#import xlrd
import openpyxl
# =====
#
# IMPORTS
#
# =====
import argparse
import codecs
import csv
import locale  # for converting strings with commas into floats
import os
import re
import sys
from collections import defaultdict

# import xlrd
import openpyxl

# Simulate an "include billing_common.py".
SCRIPT_DIR = os.path.dirname(os.path.realpath(__file__))
exec(compile(open(os.path.join(SCRIPT_DIR, "billing_common.py"), "rb").read(), os.path.join(SCRIPT_DIR, "billing_common.py"), 'exec'))

#=====
#
# CONSTANTS
#
#=====
# From billing_common.py
global BILLING_NOTIFS_SHEET_COLUMNS
global BILLING_AGGREG_SHEET_COLUMNS
global BILLING_DETAILS_PREFIX
global BILLING_NOTIFS_PREFIX
global GOOGLE_INVOICE_PREFIX
global ILAB_EXPORT_PREFIX
global CONSULTING_HOURS_FREE
global CONSULTING_TRAVEL_RATE_DISCOUNT
global ACCOUNT_PREFIXES
global SUBDIR_RAWDATA
global SUBDIR_EXPORTS
global BASE_STORAGE_SIZE

# Default headers for the ilab Export CSV file (if not read in from iLab template file).
DEFAULT_CSV_HEADERS = ['service_id','note','service_quantity','purchased_on',
                       'service_request_id','owner_email','pi_email']

#=====
#
# GLOBALS
#
#=====

#
# These globals are data structures read in from BillingConfig workbook.
#

# List of pi_tags.
pi_tag_list = []

# Mapping from usernames to list of [date, pi_tag].
username_to_pi_tag_dates = defaultdict(list)

# Mapping from usernames to a list of [email, full_name].
username_to_user_details = defaultdict(list)

# Mapping from pi_tags to list of [first_name, last_name, email].
pi_tag_to_names_email = defaultdict(list)

# Mapping from pi_tags to iLab service request IDs (1-to-1 mapping).
pi_tag_to_ilab_service_req_id = dict()

# Mapping from accounts to list of [pi_tag, %age].
account_to_pi_tag_pctages = defaultdict(list)

# Mapping from folders to list of [pi_tag, %age].
folder_to_pi_tag_pctages = defaultdict(list)

#
# These globals are data structures used to write the BillingNotification workbooks.
#

# Mapping from pi_tag to list of [folder, size, %age].
pi_tag_to_folder_sizes = defaultdict(list)

# Mapping from pi_tag to list of [account, list of [username, cpu_core_hrs, %age]].
pi_tag_to_account_username_cpus = defaultdict(list)

# Mapping from pi_tag to list of [date, username, job_name, account, cpu_core_hrs, jobID, %age].
pi_tag_to_job_details = defaultdict(list)

# Mapping from pi_tag to list of [username, date_added, date_removed, %age].
pi_tag_to_user_details = defaultdict(list)

# Mapping from pi_tag to string for their cluster service level ('Full', 'Free', 'None').
pi_tag_to_service_level = dict()

# Mapping from pi_tag to string for their affiliate status ('Stanford', 'Affiliate', 'External').
pi_tag_to_affiliation = dict()

# Mapping from pi_tag to set of (cloud account, %age) tuples.
global pi_tag_to_cloud_account_pctages
pi_tag_to_cloud_account_pctages = defaultdict(set)

# Mapping from cloud account to set of cloud project IDs (several per project possible in this set).
cloud_account_to_cloud_projects = defaultdict(set)

# Mapping from cloud account to cloud account name
cloud_account_to_account_names = dict()

# Mapping from (cloud project ID, cloud account) to lists of (platform, account, description, dates, quantity, UOM, charge) tuples.
cloud_project_account_to_cloud_details = defaultdict(list)

# Mapping from (cloud project ID, cloud account) to total charge.
cloud_project_account_to_total_charges = defaultdict(float)

# Mapping from cloud project number to cloud project ID (1-to-1 mapping).
cloud_projnum_to_cloud_project = dict()

# Mapping from cloud project ID to cloud project name (1-to-1 mapping).
cloud_projid_to_cloud_projname = dict()


# Mapping from pi_tag to list of (date, summary, hours, cumul_hours)
consulting_details = defaultdict(list)


# Set locale to be US english for converting strings with commas into floats.
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

#=====
#
# FUNCTIONS
#
#=====
# From billing_common.py
global from_timestamp_to_excel_date
global from_excel_date_to_timestamp
global from_timestamp_to_date_string
global from_excel_date_to_date_string
global from_ymd_date_to_timestamp
global from_date_string_to_timestamp
global from_timestamp_to_datetime
global from_datetime_to_timestamp
global from_datetime_to_date_string
global sheet_get_named_column
global filter_by_dates
global argparse_get_parent_parser
global argparse_get_year_month
global argparse_get_billingroot_billingconfig
global get_subdirectory
global rowcol_to_a1_cell

# This function scans the username_to_pi_tag_dates dict to create a list of [pi_tag, %age]s
# for the PIs that the given user was working for on the given date.
def get_pi_tags_for_username_by_date(username, date_timestamp):

    # Add PI Tag to the list if the given date is after date_added, but before date_removed.

    pi_tag_list = []

    pi_tag_dates = username_to_pi_tag_dates.get(username)
    if pi_tag_dates is not None:

        date_excel = from_timestamp_to_datetime(date_timestamp)

        for (pi_tag, date_added, date_removed, pctage) in pi_tag_dates:
            if date_added <= date_excel and (date_removed == '' or date_removed is None or date_removed >= date_excel):
                pi_tag_list.append([pi_tag, pctage])

    return pi_tag_list

#
# Reads a subtable from the CSVFile file-object, which is all the lines
# between blank lines.
#
def get_google_invoice_csv_subtable_lines(csvfile_obj):

    subtable = []

    line = csvfile_obj.readline()
    while not line.startswith(',') and line != '' and line != '\n':
        subtable.append(line)
        line = csvfile_obj.readline()

    return subtable


# Creates all the data structures used to write the BillingNotification workbook.
# The overall goal is to mimic the tables of the notification sheets so that
# to build the table, all that is needed is to print out one of these data structures.
def build_global_data(billing_config_wkbk, begin_month_timestamp, end_month_timestamp, read_cloud_data):

    pis_sheet      = billing_config_wkbk["PIs"]
    folders_sheet  = billing_config_wkbk["Folders"]
    users_sheet    = billing_config_wkbk["Users"]
    accounts_sheet = billing_config_wkbk["Accounts"]

    begin_month_exceldate = from_timestamp_to_excel_date(begin_month_timestamp)
    end_month_exceldate   = from_timestamp_to_excel_date(end_month_timestamp)

    begin_month_datetime = from_timestamp_to_datetime(begin_month_timestamp)
    end_month_datetime   = from_timestamp_to_datetime(end_month_timestamp)

    #
    # Create list of pi_tags.
    #
    global pi_tag_list

    pi_tag_list = list(sheet_get_named_column(pis_sheet, "PI Tag"))
    # Remove all empty cells from the end of the pi_tag_list
    while pi_tag_list[-1] is None:
        pi_tag_list = pi_tag_list[:-1]

    #
    # Create mapping from pi_tag to a list of PI name and email.
    #
    global pi_tag_to_names_email

    pi_first_names = sheet_get_named_column(pis_sheet, "PI First Name")
    pi_last_names  = sheet_get_named_column(pis_sheet, "PI Last Name")
    pi_emails      = sheet_get_named_column(pis_sheet, "PI Email")
    owner_emails   = sheet_get_named_column(pis_sheet, "iLab Service Request Owner")

    pi_details_list = list(zip(pi_first_names, pi_last_names, pi_emails, owner_emails))

    pi_tag_to_names_email = dict(list(zip(pi_tag_list, pi_details_list)))

    #
    # Create mapping from pi_tag to iLab Service Request ID.
    #
    global pi_tag_to_ilab_service_req_id

    pi_ilab_ids = sheet_get_named_column(pis_sheet, "iLab Service Request ID")

    pi_tag_to_ilab_service_req_id = dict(list(zip(pi_tag_list, pi_ilab_ids)))

    # Organize data from the Cloud sheet, if present.
    if read_cloud_data:
        cloud_sheet = billing_config_wkbk["Cloud Accounts"]

        #
        # Create mapping from pi_tag to (cloud account, %age) tuples from the BillingConfig PIs sheet.
        # Create mapping from cloud account to account names
        #
        global pi_tag_to_cloud_account_pctages
        global cloud_account_to_account_names

        cloud_platforms = sheet_get_named_column(cloud_sheet, "Platform")
        cloud_pi_tags = sheet_get_named_column(cloud_sheet, "PI Tag")
        cloud_accounts = sheet_get_named_column(cloud_sheet, "Billing Account Number")
        cloud_account_names = sheet_get_named_column(cloud_sheet, "Billing Account Name")
        cloud_pctages = sheet_get_named_column(cloud_sheet, "%age")

        cloud_dates_added = sheet_get_named_column(cloud_sheet, "Date Added")
        cloud_dates_remvd = sheet_get_named_column(cloud_sheet, "Date Removed")

        cloud_rows = filter_by_dates(list(zip(cloud_platforms, cloud_pi_tags,
                                              cloud_accounts, cloud_account_names, cloud_pctages)),
                                     list(zip(cloud_dates_added, cloud_dates_remvd)),
                                     begin_month_datetime, end_month_datetime)

        # for (pi_tag, project, projnum, projid, account, pctage) in cloud_rows:
        for (platform, pi_tag, account, acct_name, pctage) in cloud_rows:

            # Only Google Cloud is supported by automated billing (for now)
            if platform != "Google": continue

            # Associate the project name and percentage to be charged with the pi_tag.
            pi_tag_to_cloud_account_pctages[pi_tag].add((account, pctage))

            # Associate the account name with the account
            cloud_account_to_account_names[account] = acct_name

    #
    # Create mapping from pi_tags to a string denoting service level on cluster.
    #
    global pi_tag_to_service_level

    levels_column = sheet_get_named_column(pis_sheet,"Cluster?")

    pi_tag_to_service_level = dict(list(zip(pi_tag_list, levels_column)))

    #
    # Create mapping from pi_tags to a string denoting affiliation (Stanford/Affiliate/External).
    #
    global pi_tag_to_affiliation

    affiliation_column = sheet_get_named_column(pis_sheet, "Affiliation")

    pi_tag_to_affiliation = dict(list(zip(pi_tag_list, affiliation_column)))

    #
    # Filter pi_tag_list for PIs active in the current month.
    #
    pi_dates_added   = sheet_get_named_column(pis_sheet, "Date Added")
    pi_dates_removed = sheet_get_named_column(pis_sheet, "Date Removed")

    pi_tags_and_dates_added = list(zip(pi_tag_list, pi_dates_added, pi_dates_removed))

    for (pi_tag, date_added, date_removed) in pi_tags_and_dates_added:

        # Convert the Excel dates to timestamps.
        if date_added is None: continue
        date_added_timestamp = from_datetime_to_timestamp(date_added)
        if date_removed != '' and date_removed is not None:
            date_removed_timestamp = from_datetime_to_timestamp(date_removed)
        else:
            date_removed_timestamp = end_month_timestamp + 1  # Not in this month.

        # If the date added is AFTER the end of this month, or
        #  the date removed is BEFORE the beginning of this month,
        # then remove the pi_tag from the list.
        if date_added_timestamp >= end_month_timestamp:

            print(" *** Ignoring PI %s: added after this month on %s" % (pi_tag_to_names_email[pi_tag][1], from_datetime_to_date_string(date_added)), file=sys.stderr)
            pi_tag_list.remove(pi_tag)

        elif date_removed_timestamp < begin_month_timestamp:

            print(" *** Ignoring PI %s: removed before this month on %s" % (pi_tag_to_names_email[pi_tag][1], from_datetime_to_date_string(date_removed)), file=sys.stderr)
            pi_tag_list.remove(pi_tag)

    #
    # Create mapping from usernames to a list of user details.
    #
    global username_to_user_details

    usernames  = sheet_get_named_column(users_sheet, "Username")
    emails     = sheet_get_named_column(users_sheet, "Email")
    full_names = sheet_get_named_column(users_sheet, "Full Name")

    username_details_rows = list(zip(usernames, emails, full_names))

    for (username, email, full_name) in username_details_rows:
        username_to_user_details[username] = [email, full_name]

    #
    # Create mapping from usernames to a list of pi_tag/dates.
    #
    global username_to_pi_tag_dates

    pi_tags       = sheet_get_named_column(users_sheet, "PI Tag")
    dates_added   = sheet_get_named_column(users_sheet, "Date Added")
    dates_removed = sheet_get_named_column(users_sheet, "Date Removed")
    pctages       = sheet_get_named_column(users_sheet, "%age")

    username_rows = list(zip(usernames, pi_tags, dates_added, dates_removed, pctages))

    for (username, pi_tag, date_added, date_removed, pctage) in username_rows:
        username_to_pi_tag_dates[username].append([pi_tag, date_added, date_removed, pctage])

    #
    # Create mapping from pi_tags to a list of [username, date_added, date_removed]
    #
    global pi_tag_to_user_details

    for username in username_to_pi_tag_dates:

        pi_tag_date_list = username_to_pi_tag_dates[username]

        for (pi_tag, date_added, date_removed, pctage) in pi_tag_date_list:
            pi_tag_to_user_details[pi_tag].append([username, date_added, date_removed, pctage])

    #
    # Create mapping from account to list of pi_tags and %ages.
    #
    global account_to_pi_tag_pctages

    accounts = sheet_get_named_column(accounts_sheet, "Account")
    pi_tags  = sheet_get_named_column(accounts_sheet, "PI Tag")
    pctages  = sheet_get_named_column(accounts_sheet, "%age")

    dates_added   = sheet_get_named_column(accounts_sheet, "Date Added")
    dates_removed = sheet_get_named_column(accounts_sheet, "Date Removed")

    account_rows = filter_by_dates(list(zip(accounts, pi_tags, pctages)), list(zip(dates_added, dates_removed)),
                                   #begin_month_exceldate, end_month_exceldate)
                                   begin_month_datetime, end_month_datetime)

    for (account, pi_tag, pctage) in account_rows:
        account_to_pi_tag_pctages[account].append([pi_tag, pctage])

    # Add pi_tags prefixed by ACCOUNT_PREFIXES to list of accounts for PI.
    for pi_tag in pi_tag_list:
        account_to_pi_tag_pctages[pi_tag].append([pi_tag, 1.0])

        for prefix in ACCOUNT_PREFIXES:
            account_to_pi_tag_pctages["%s_%s" % (prefix,pi_tag)].append([pi_tag, 1.0])

    #
    # Create mapping from folder to list of pi_tags and %ages.
    #
    global folder_to_pi_tag_pctages

    # Get the Folders from PI Sheet
    folders = sheet_get_named_column(pis_sheet, "PI Folder")
    pi_tags = sheet_get_named_column(pis_sheet, "PI Tag")
    pctages = [1.0] * len(folders)

    dates_added   = sheet_get_named_column(pis_sheet, "Date Added")
    dates_removed = sheet_get_named_column(pis_sheet, "Date Removed")

    # Add the Folders from Folder sheet
    folders += sheet_get_named_column(folders_sheet, "Folder")
    pi_tags += sheet_get_named_column(folders_sheet, "PI Tag")
    pctages += sheet_get_named_column(folders_sheet, "%age")

    dates_added   += sheet_get_named_column(folders_sheet, "Date Added")
    dates_removed += sheet_get_named_column(folders_sheet, "Date Removed")

    folder_rows = filter_by_dates(list(zip(folders, pi_tags, pctages)), list(zip(dates_added, dates_removed)),
                                  #begin_month_exceldate, end_month_exceldate)
                                  begin_month_datetime, end_month_datetime)

    for (folder, pi_tag, pctage) in folder_rows:

        # Account for multiple folders separated by commas.
        pi_folder_list = folder.split(',')

        for pi_folder in pi_folder_list:
            folder_to_pi_tag_pctages[pi_folder].append([pi_tag, pctage])


# Reads the particular rate requested from the Rates sheet of the BillingConfig workbook.
def get_rate(wkbk, rate_type):

    rates_sheet = wkbk["Rates"]

    types   = sheet_get_named_column(rates_sheet, 'Type')
    amounts = sheet_get_named_column(rates_sheet, 'Amount')

    for (type, amount) in zip(types, amounts):
        if type == rate_type:
            return amount
    else:
        return None


def get_rate_a1_cell(wkbk, rate_string):

    rates_sheet = wkbk["Rates"]

    header_row = rates_sheet.iter_cols(min_row=1, max_row=1, values_only=True)

    # Find the column numbers for 'Type' and 'Amount'.
    type_col = -1
    amt_col = -1
    idx = 1
    for col_name in header_row:
        if col_name[0] == 'Type':
            type_col = idx
            if amt_col != -1: break  # Leave loop if we have both
        elif col_name[0] == 'Amount':
            amt_col = idx
            if type_col != -1: break  # Leave loop if we have both
        idx += 1

    if type_col == -1 or amt_col == -1:
        print("get_rate_a1_cell: Can't find Type/Amount headers (%d, %d)" % (type_col, amt_col), file=sys.stderr)
        return None

    # Get column of 'Types'.
    types = rates_sheet.iter_rows(min_row=2, min_col=type_col, max_col=type_col, values_only=True)

    # When you find the row with rate_string, return the Amount col and this row.
    idx = 2
    for row in types:
        for col in row:
            if col == rate_string:
                return 'Rates!%s' % rowcol_to_a1_cell(idx + 1, amt_col, True, True)
        idx += 1
    else:
        return 0.0


def get_rate_service_id(wkbk, rate_string):

    rates_sheet = wkbk["Rates"]

    types    = sheet_get_named_column(rates_sheet, 'Type')
    serv_ids = sheet_get_named_column(rates_sheet, 'iLab Service ID')

    for (type, serv_id) in zip(types, serv_ids):
        if type == rate_string:
            return serv_id
    else:
        return None


def get_rate_data_from_string(billing_config_wkbk, service_str, tier_str, subservice_str, affiliation_str):

    rate_string = service_str  # To start, will be appended to below

    if service_str == "Local HPC Storage" or service_str == "Local Computing":
        # Add the tier string
        rate_string += " (%s Tier)" % tier_str.capitalize()

    # If there is a subservice string, add that to rate string
    if subservice_str is not None and subservice_str != "":
        rate_string += ": %s" % subservice_str

    # Finish rate string with the affiliation string
    rate_string += " - %s" % affiliation_str.capitalize()

    # Look up rate amount, rate cell from rate_string
    rate_amount     = get_rate(billing_config_wkbk, rate_string)
    rate_a1_cell    = get_rate_a1_cell(billing_config_wkbk, rate_string)
    rate_service_id = get_rate_service_id(billing_config_wkbk, rate_string)

    return rate_amount, rate_a1_cell, rate_service_id


# Reads the Storage sheet of the BillingDetails workbook given, and populates
# the pi_tag_to_folder_sizes dict with the folder measurements for each PI.
def read_storage_sheet(wkbk):

    global pi_tag_to_folder_sizes

    storage_sheet = wkbk["Storage"]

    for (date, timestamp, folder, size, used, inodes_quota, inodes_used) in storage_sheet.iter_rows(min_row=2, values_only=True):

        # List of [pi_tag, %age] pairs.
        pi_tag_pctages = folder_to_pi_tag_pctages[folder]

        for (pi_tag, pctage) in pi_tag_pctages:
            pi_tag_to_folder_sizes[pi_tag].append([folder, size, pctage])


# Reads the Computing sheet of the BillingDetails workbook given, and populates
# the account_to_pi_tag_cpus, pi_tag_to_account_username_cpus, and pi_tag_to_job_details dicts.
def read_computing_sheet(wkbk):

    global pi_tag_to_job_details

    computing_sheet = wkbk["Computing"]

    if args.cpu_time_unit == 'cpu-hours':
        cpu_time_denom = 3600.0
    elif args.cpu_time_unit == 'cpu-days':
        cpu_time_denom = 86400.0
    else:
        print("Arg 'cpu_time_unit' has unknown value {args.cpu_time_unit", file=sys.stderr)

    sheet_number = 1

    while True:

        for (job_date, job_timestamp, job_username, job_name, account, node, cores, wallclock, jobID) in \
                computing_sheet.iter_rows(min_row=2, values_only=True):

            # Calculate CPU-core units for job.
            cpu_core_time = cores * wallclock / cpu_time_denom   # wallclock is in seconds.

            # Rename this variable for easier understanding.
            account = account.lower()

            if account != '':
                job_pi_tag_pctage_list = account_to_pi_tag_pctages[account]
            else:
                # No account means credit the job to the user's lab.
                job_pi_tag_pctage_list = get_pi_tags_for_username_by_date(job_username, job_timestamp)

            if len(job_pi_tag_pctage_list) == 0:
                print("   *** No PI associated with job ID %d, pi_tag %s, account %s" % (jobID, pi_tag, account))
                continue

            # Distribute this job's CPU-units amongst pi_tags by %ages.
            for (pi_tag, pctage) in job_pi_tag_pctage_list:

                # This list is [account, list of [username, cpu_core_time, %age]].
                account_username_cpu_list = pi_tag_to_account_username_cpus.get(pi_tag)

                # If pi_tag has an existing list of account/username/CPUs:
                if account_username_cpu_list is not None:

                    # Find if account for job is in list of account/CPUs for this pi_tag.
                    for pi_username_cpu_pctage_list in account_username_cpu_list:

                        (pi_account, pi_username_cpu_pctage_list) = pi_username_cpu_pctage_list

                        # If the account we are looking at is the one from the present job:
                        if pi_account == account:

                            # Find job username in list for account:
                            for username_cpu in pi_username_cpu_pctage_list:
                                if job_username == username_cpu[0]:
                                    username_cpu[1] += cpu_core_time
                                    break
                            else:
                                pi_username_cpu_pctage_list.append([job_username, cpu_core_time, pctage])

                            # Leave account_username_cpu_list loop.
                            break

                    else:
                        # No matching account in pi_tag list -- add a new one to the list.
                        account_username_cpu_list.append([account, [[job_username, cpu_core_time, pctage]]])

                # Else start a new account/CPUs list for the pi_tag.
                else:
                    pi_tag_to_account_username_cpus[pi_tag] = [[account, [[job_username, cpu_core_time, pctage]]]]

                #
                # Save job details for pi_tag.
                #
                new_job_details = [job_date, job_username, job_name, account, node, cpu_core_time, jobID, pctage]
                pi_tag_to_job_details[pi_tag].append(new_job_details)

        sheet_number += 1

        try:
            computing_sheet = wkbk["Computing %d" % sheet_number]
        except:
            break  # No more computing sheets: exit the while True loop.


# Read the Cloud sheet from the BillingDetails workbook.
def read_cloud_sheet(wkbk):

    cloud_sheet = wkbk["Cloud"]

    for (platform, account, project, description, dates, quantity, uom, charge) in \
            cloud_sheet.iter_rows(min_row=2, values_only=True):

        # If project is of the form "<project name>(<project-id>)" or "<project name>[<project-id>]", get the "<project-id>".
        if project is not None:
            project_re = re.search("[(\[]([a-z0-9-:.]+)[\])]", project)
            if project_re is not None:
                project = project_re.group(1)
            else:
                pass  # If no parens, use the original project name.

        # Save the project that the account line item is for.
        cloud_account_to_cloud_projects[account].add(project)

        # Save the cloud item in a list of charges for that PI.
        cloud_project_account_to_cloud_details[(project, account)].append((platform, description, dates, quantity, uom, charge))

        # Accumulate the total cost of a project.
        cloud_project_account_to_total_charges[(project, account)] += float(charge)


def read_google_invoice(google_invoice_csv_file):

    ###
    # Read the Google Invoice CSV File
    ###

    # Google Invoice CSV files are Unicode with BOM.
    google_invoice_csv_file_obj = codecs.open(google_invoice_csv_file, 'rU', encoding='utf-8-sig')

    #  Read the header subtable
    google_invoice_header_subtable = get_google_invoice_csv_subtable_lines(google_invoice_csv_file_obj)

    google_invoice_header_csvreader = csv.DictReader(google_invoice_header_subtable, fieldnames=['key', 'value'])

    for row in google_invoice_header_csvreader:

        #   Extract invoice date from "Issue Date".
        if row['key'] == 'Issue date':
            google_invoice_issue_date = row['value']
        #   Extract the "Amount Due" value.
        elif row['key'] == 'Amount due':
            google_invoice_amount_due = locale.atof(row['value'])

    print("  Amount due: $%0.2f" % (google_invoice_amount_due), file=sys.stderr)

    # Accumulate the total amount of charges while processing each line,
    #  to compare with total amount in header.
    google_invoice_total_amount = 0.0

    #  While there are still more subtables...
    while True:

        #   Read subtable.
        google_invoice_subtable = get_google_invoice_csv_subtable_lines(google_invoice_csv_file_obj)

        #   No more subtables?!  Let's get out of here!
        if len(google_invoice_subtable) == 0:
            break

        #   Create CSVReader from subtable
        google_invoice_subtable_csvreader = csv.DictReader(google_invoice_subtable)

        #   Foreach row in CSVReader
        for row_dict in google_invoice_subtable_csvreader:

            #     Accumulate total charges.
            amount = locale.atof(row_dict['Amount'])
            google_invoice_total_amount += amount

            google_account = row_dict['Order']

            #     Construct note for ilab entry.
            google_platform = 'Google Cloud Platform, Firebase, and APIs'
            google_project = row_dict['Source']
            google_item    = row_dict['Description']
            google_quantity = row_dict['Quantity']
            google_uom     = row_dict['UOM']
            google_dates   = row_dict['Interval']

            # Save the cloud details with the appropriate PI.
            cloud_project_account_to_cloud_details[(google_project, google_account)].append((google_platform, google_item, google_dates,
                                                                                             google_quantity, google_uom, amount))

    # Compare total charges to "Amount Due".
    if abs(google_invoice_total_amount - google_invoice_amount_due) >= 0.01:  # Ignore differences less than a penny.
        print("  WARNING: Accumulated amounts do not equal amount due: ($%.2f != $%.2f)" % (google_invoice_total_amount,
                                                                                            google_invoice_amount_due), file=sys.stderr)
    else:
        print("  VERIFIED: Sum of individual transactions equals Amount due.", file=sys.stderr)


#
# Read in the Consulting sheet.
#
# It fills in the dict consulting_details.
#
def read_consulting_sheet(wkbk):

    #consulting_sheet = wkbk.sheet_by_name("Consulting")
    consulting_sheet = wkbk["Consulting"]

    #for row in range(1, consulting_sheet.nrows):
    #(date, pi_tag, hours, travel_hours, participants, clients, summary, notes, cumul_hours) = consulting_sheet.row_values(row)

    for (date, pi_tag, hours, travel_hours, participants, clients, summary, notes, cumul_hours) in \
            consulting_sheet.iter_rows(min_row=2, values_only=True):

        if travel_hours is None: travel_hours = 0

        # Save the consulting item in a list of charges for that PI.
        consulting_details[pi_tag].append((date, summary, clients, float(hours), float(travel_hours), float(cumul_hours)))


#
# Digest cluster data and output Cluster iLab file.
#
def process_cluster_data():

    # Read in its Storage sheet.
    print("Reading storage sheet.")
    read_storage_sheet(billing_details_wkbk)

    # Read in its Computing sheet.
    print("Reading computing sheet.")
    read_computing_sheet(billing_details_wkbk)


def open_ilab_output_dictwriter(subdir, suffix):

    ###
    #
    # Open an iLab CSV file for writing out.
    #
    ###
    ilab_export_csv_filename = "%s-%s.%s-%02d.csv" % (ILAB_EXPORT_PREFIX, suffix, year, month)
    ilab_export_csv_pathname = os.path.join(subdir, ilab_export_csv_filename)

    ilab_export_csv_file = open(ilab_export_csv_pathname, "w")

    ilab_export_csv_dictwriter = csv.DictWriter(ilab_export_csv_file, fieldnames=ilab_csv_headers)

    ilab_export_csv_dictwriter.writeheader()

    return ilab_export_csv_dictwriter

#
# Digest cloud data and output Cloud iLab file.
#
def process_cloud_data():

    # Read in Cloud data from Google Invoice, if given as argument.
    if args.google_invoice_csv is not None:

        ###
        # Read in Google Cloud Invoice data, ignoring data from BillingDetails.
        ###
        print("Reading Google Invoice.")
        read_google_invoice(google_invoice_csv)

    # Read in the Cloud sheet from the BillingDetails file, if present.
    elif "Cloud" in billing_details_wkbk.sheetnames:

        print("Reading cloud sheet.")

        read_cloud_sheet(billing_details_wkbk)

    else:
        print("No Cloud sheet in BillingDetails nor Google Invoice file...skipping")
        return


#
# Digest Consulting data and output Consulting iLab file.
#
def process_consulting_data():

    # Read in its Consulting sheet.
    if "Consulting" in billing_details_wkbk.sheetnames:
        print("Reading consulting sheet.")
        read_consulting_sheet(billing_details_wkbk)
    else:
        print("No consulting sheet in BillingDetails: skipping")
        return


#
# Generates the iLab Cluster Storage CSV entries for a particular pi_tag.
#
# It uses dict pi_tag_to_folder_sizes.
#
def output_ilab_csv_data_for_cluster_storage(csv_dictwriter, pi_tag, service_req_id, storage_base_service_id, storage_addl_service_id,
                                             begin_month_timestamp, end_month_timestamp):

    purchased_on_date = from_timestamp_to_date_string(end_month_timestamp-1)  # Last date of billing period.

    ###
    #
    # STORAGE Subtable
    #
    ###
    output_storage_p = False  # Did any lines get output?
    total_storage_sizes = 0.0
    if storage_base_service_id is not None:

        for (folder, size, pctage) in pi_tag_to_folder_sizes[pi_tag]:

            if folder == '/labs/%s' % pi_tag and size >= BASE_STORAGE_SIZE:
                # Note format: <folder>
                note = "%s" % (folder)
                output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, storage_base_service_id, note,1)

                total_storage_sizes += BASE_STORAGE_SIZE

                size -= BASE_STORAGE_SIZE

                output_storage_p = True

            if size > 0.0:
                # Note format: <folder> [<pct>%, if not 0%]
                note = "%s" % (folder)

                if 0.0 < pctage < 1.0:
                    note += " [%d%%]" % (pctage * 100)

                quantity = size * pctage

                if quantity > 0.0:
                    output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, storage_addl_service_id, note, quantity)

                    total_storage_sizes += size

                    output_storage_p = True

    return output_storage_p


#
# Generates the iLab Cluster Computing CSV entries for a particular pi_tag.
#
# It uses dicts pi_tag_to_username_cpus, and pi_tag_to_account_cpus.
#
def output_ilab_csv_data_for_cluster_compute(csv_dictwriter, pi_tag, service_req_id, lab_computing_service_id, full_computing_service_id,
                                             begin_month_timestamp, end_month_timestamp):

    purchased_on_date = from_timestamp_to_date_string(end_month_timestamp-1)  # Last date of billing period.

    ###
    #
    # COMPUTING Subtable
    #
    ###

    # Loop over pi_tag_to_account_username_cpus for account/username combos.
    account_username_cpus_list = pi_tag_to_account_username_cpus.get(pi_tag)

    output_compute_p = False   # Were any lines written out?
    if account_username_cpus_list is not None:

        for (account, username_cpu_pctage_list) in account_username_cpus_list:

            if len(username_cpu_pctage_list) > 0:

                for (username, cpu_core_hrs, pctage) in username_cpu_pctage_list:

                    fullname = username_to_user_details[username][1]

                    # Note format: <user-name> (<user-ID>) [<pct>%, if not 0%]
                    note = "Account: %s - User: %s (%s)" % (account, fullname, username)

                    if 0.0 < pctage < 1.0:
                        note += " [%d%%]" % (pctage * 100)

                    quantity = cpu_core_hrs * pctage

                    if quantity > 0.0:
                        if lab_computing_service_id is not None:
                            output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, lab_computing_service_id,
                                                     note, quantity)
                            output_compute_p = True
                        # Lab is in Free Tier, and we can charge them if someone outside the lab ran the job.
                        else:
                            pi_tags_for_username = get_pi_tags_for_username_by_date(username, begin_month_timestamp)

                            # If the user is not within the lab membership, then use the full tier service ID.
                            if pi_tag not in [pi_pct[0] for pi_pct in pi_tags_for_username]:
                                output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, full_computing_service_id, note, quantity)
                                output_compute_p = True
                            else:
                                print("  *** In Free Tier Lab %s, lab member %s ran billable jobs (%f)." % (pi_tag, username, quantity), file=sys.stderr)

            else:
                # No users for this PI.
                pass

    return output_compute_p

#
# Generates the iLab Cloud CSV entries for a particular pi_tag.
#
# It uses dicts pi_tag_to_cloud_account_pctages and cloud_project_account_to_cloud_details.
#
def output_ilab_csv_data_for_cloud(csv_dictwriter, pi_tag, service_req_id, cloud_service_id,
                                   begin_month_timestamp, end_month_timestamp):

    purchased_on_date = from_timestamp_to_date_string(end_month_timestamp-1) # Last date of billing period.

    # Get PI Last name for some situations below.
    (_, pi_last_name, _, _) = pi_tag_to_names_email[pi_tag]

    # Get list of (account, %ages) tuples for given PI.
    output_cloud_p = False  # Were any lines written out?
    for (account, pctage) in pi_tag_to_cloud_account_pctages[pi_tag]:

        if pctage == 0.0: continue

        account_name = cloud_account_to_account_names[account]
        if account_name is None or account_name == "":
            account_name = account

        for project_id in cloud_account_to_cloud_projects[account]:

            # Get list of cloud items to charge PI for.
            cloud_details = cloud_project_account_to_cloud_details[(project_id, account)]

            # Get name for project ID.
            project_name = cloud_projid_to_cloud_projname.get(project_id)
            if project_name is None:
                project_name = project_id

            if not args.break_out_cloud and len(cloud_details) > 0:

                # Generate a single transaction of all the transactions within the project.
                total_amount_for_project = 0

                # Add up all the charges for that project within the cloud details.
                for (platform, description, dates, quantity, uom, amount) in cloud_details:
                    total_amount_for_project += amount

                pi_amount = total_amount_for_project * pctage
                if pi_amount == 0.0: continue

                # Create a note for the rolled-up transactions.
                if project_name is not None:
                    note = "Google :: Charges for Project '%s' (%s)" % (project_name, project_id)
                else:
                    note = "Google :: Misc charges/credits for %s " % (pi_last_name)

                if pctage < 1.0:
                    note += " [%d%%]" % (pctage * 100)

                # Output the single transaction for the project.
                output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, cloud_service_id, note,
                                         pi_amount)

                output_cloud_p = True
            else:
                for (platform, description, dates, quantity, uom, amount) in cloud_details:

                    # If the quantity is given, make a string of it and its unit-of-measure.
                    if quantity != '':
                        quantity_str = " @ %s %s" % (quantity, uom)
                    else:
                        quantity_str = ''

                    if project_name != '':
                        proj_str = "%s (%s)" % (project_name, project_id)
                    else:
                        proj_str = 'Misc charges/credits for %s' % pi_last_name

                    note = "Google :: %s : %s%s" % (proj_str, description, quantity_str)

                    if pctage < 1.0:
                        note += " [%d%%]" % (pctage * 100)

                    # Calculate the amount to charge the PI based on their percentage.
                    pi_amount = amount * pctage
                    if pi_amount == 0.0: continue

                    # Write out the iLab export line.
                    output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, cloud_service_id, note, pi_amount)

                    output_cloud_p = True

    return output_cloud_p


def output_ilab_csv_data_for_consulting(csv_dictwriter, pi_tag, service_req_id, consulting_free_service_id, consulting_paid_service_id,
                                        begin_month_timestamp, end_month_timestamp):

    output_consulting_p = False
    for (date, summary, client, hours, travel_hours, cumul_hours) in consulting_details[pi_tag]:

        date_timestamp    = from_datetime_to_timestamp(date)
        purchased_on_date = from_datetime_to_date_string(date)

        # If this transaction occurred within this month:
        if begin_month_timestamp <= date_timestamp < end_month_timestamp:

            #
            # Calculate the number of free hours and paid hours in this transaction.
            #
            start_hours_used = cumul_hours - hours - travel_hours

            if start_hours_used < CONSULTING_HOURS_FREE:
                free_hours_remaining = CONSULTING_HOURS_FREE - start_hours_used
            else:
                free_hours_remaining = 0

            if hours < free_hours_remaining:
                free_hours_used = hours
            else:
                free_hours_used = free_hours_remaining

            paid_hours_used = hours - free_hours_used + (travel_hours * CONSULTING_TRAVEL_RATE_DISCOUNT)

            # Write out the iLab export line for the free hours used.
            if free_hours_used > 0:
                output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, consulting_free_service_id,
                                         "%s [%s]" % (summary, client), free_hours_used)
                output_consulting_p = True

            # Write out the iLab export line for the paid hours used.
            if paid_hours_used > 0:
                output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, purchased_on_date, consulting_paid_service_id,
                                         "%s [%s]" % (summary, client), paid_hours_used)
                output_consulting_p = True

    return output_consulting_p


def output_ilab_csv_data_row(csv_dictwriter, pi_tag, service_req_id, end_month_string, service_id, note, amount):

    # Confirm we have a real service ID and not the "Free" of Free Tier
    if not re.match("[0-9]+", str(service_req_id)):
        print("\n   Free Tier PI {} has charge for {}...ignoring".format(pi_tag, note, file=sys.stderr))
        return False 

    # Create a dictionary to be written out as CSV.
    csv_dict = dict()
    # If there is an 'iLab Owner Email' available, use that, o/w, use the PI email.
    if pi_tag_to_names_email[pi_tag][3] != '' and pi_tag_to_names_email[pi_tag][3] is not None:
        csv_dict['owner_email'] = pi_tag_to_names_email[pi_tag][3]
    else:
        csv_dict['owner_email'] = pi_tag_to_names_email[pi_tag][2]
    csv_dict['pi_email']     = pi_tag_to_names_email[pi_tag][2]
    csv_dict['service_request_id'] = int(service_req_id)
    csv_dict['purchased_on'] = end_month_string  # Last date of billing period.
    csv_dict['service_id'] = service_id

    csv_dict['note'] = note
    csv_dict['service_quantity'] = amount

    csv_dictwriter.writerow(csv_dict)

    return True


#=====
#
# SCRIPT BODY
#
#=====

parser = argparse.ArgumentParser(parents=[argparse_get_parent_parser()])

parser.add_argument("-D","--billing_details_file",
                    default=None,
                    help='The BillingDetails file')
parser.add_argument("-g", "--google_invoice_csv",
                    default=None,
                    help="The Google Invoice CSV file")
parser.add_argument("-t", "--ilab_template",
                    default=None,
                    help='The iLab export file template [default = None]')
parser.add_argument("-c", "--skip_cluster", action="store_true",
                    default=False,
                    help="Don't output cluster iLab files. [default = False]")
parser.add_argument("-S", "--skip_cluster_storage", action="store_true",
                    default=False,
                    help="Don't output cluster storage iLab file. [default = False]")
parser.add_argument("-C", "--skip_cluster_compute", action="store_true",
                    default=False,
                    help="Don't output cluster compute iLab file. [default = False]")
parser.add_argument("-l", "--skip_cloud", action="store_true",
                    default=False,
                    help="Don't output cloud iLab file. [default = False]")
parser.add_argument("-n", "--skip_consulting", action="store_true",
                    default=False,
                    help="Don't output consulting iLab file. [default = False]")
parser.add_argument( "--break_out_cloud", action="store_true",
                     default=False,
                     help="Break out individual cloud transactions. [default = False]")
parser.add_argument("--cpu_time_unit", choices=['cpu-hours', 'cpu-days'],
                    default='cpu-days',
                    help='Choose the CPU time units [default = cpu-days]')

args = parser.parse_args()

#
# Process arguments.
#

# Get year/month-related arguments
(year, month, begin_month_timestamp, end_month_timestamp) = argparse_get_year_month(args)

# Get BillingRoot and BillingConfig arguments
(billing_root, billing_config_file) = argparse_get_billingroot_billingconfig(args, year, month)

###
#
# Read the BillingConfig workbook and build input data structures.
#
###

#billing_config_wkbk = xlrd.open_workbook(billing_config_file)
billing_config_wkbk = openpyxl.load_workbook(billing_config_file)

# Build path to the input files
input_subdir = get_subdirectory(billing_root, year, month, "")

# If BillingDetails file given, use that, else look in BillingRoot.
if args.billing_details_file is not None:
    billing_details_file = args.billing_details_file
else:
    billing_details_file = os.path.join(input_subdir, "%s.%s-%02d.xlsx" % (BILLING_DETAILS_PREFIX, year, month))

# Get the absolute path for the billing_details_file.
billing_details_file = os.path.abspath(billing_details_file)

# Confirm that BillingDetails file exists.
if not os.path.exists(billing_details_file):
    print("BillingDetailsFile %s doesn't exist" % billing_details_file, file=sys.stderr)
    sys.exit(-1)

# Build path to a possible GoogleInvoice file within BillingRoot
google_input_subdir = get_subdirectory(billing_root, year, month, SUBDIR_RAWDATA)

# If Google Invoice CSV given, use that, else look in BillingRoot.
if args.google_invoice_csv is not None:
    google_invoice_csv = args.google_invoice_csv
else:
    google_invoice_filename = "%s.%d-%02d.csv" % (GOOGLE_INVOICE_PREFIX, year, month)
    google_invoice_csv = os.path.join(google_input_subdir, google_invoice_filename)

# Get absolute path for google_invoice_csv file.
google_invoice_csv = os.path.abspath(google_invoice_csv)

# Confirm that Google Invoice CSV file exists.
if not os.path.exists(google_invoice_csv):
    print("GoogleInvoice %s doesn't exist" % google_invoice_csv, file=sys.stderr)
    sys.exit(-1)

# Build a path within BillingRoot to the output directory for iLab files, creating the dir if necessary.
output_subdir = get_subdirectory(billing_root, year, month, SUBDIR_EXPORTS, create_if_nec=True)

#
# Output the state of arguments.
#
print("GENERATING ILAB EXPORT FOR %02d/%d:" % (month, year))
print("  BillingConfigFile: %s" % billing_config_file)
print("  BillingRoot: %s" % billing_root)
print("  BillingDetailsFile: %s" % billing_details_file)
print("  GoogleInvoiceCSV: %s" % google_invoice_csv)
print()

#
# Build data structures.
#
print("Building configuration data structures.")

# Determine whether we should read in Cloud data from the BillingConfig spreadsheet.
# We should if the BillingConfig spreadsheet has a Cloud sheet.
read_cloud_data = ("Cloud Accounts" in billing_config_wkbk.sheetnames)

build_global_data(billing_config_wkbk, begin_month_timestamp, end_month_timestamp, read_cloud_data)

###
#
# Read the BillingDetails workbook.
#
###

# Open the BillingDetails workbook.
print("Opening BillingDetails workbook...")
billing_details_wkbk = openpyxl.load_workbook(billing_details_file)

###
#
# Read in the iLab Export File template, if available.
#
###
if args.ilab_template is not None:

    ilab_template_file = open(args.ilab_template)
    csv_reader = csv.reader(ilab_template_file)
    ilab_csv_headers = next(csv_reader)
    ilab_template_file.close()

else:
    ilab_csv_headers = DEFAULT_CSV_HEADERS

#####
#
# Output Cluster data into iLab Cluster export file, if requested.
#
####
if billing_details_file is not None and not args.skip_cluster and \
        not (args.skip_cluster_storage and args.skip_cluster_compute):
    process_cluster_data()

    if not args.skip_cluster_storage:
        ilab_cluster_storage_export_csv_dictwriter = open_ilab_output_dictwriter(output_subdir, "Cluster_Storage")
    else:
        ilab_cluster_storage_export_csv_dictwriter = None

    if not args.skip_cluster_compute:
        ilab_cluster_compute_export_csv_dictwriter = open_ilab_output_dictwriter(output_subdir, "Cluster_Compute")
    else:
        ilab_cluster_compute_export_csv_dictwriter = None
else:
    ilab_cluster_storage_export_csv_dictwriter = None
    ilab_cluster_compute_export_csv_dictwriter = None

###
#
# Output Cloud data into iLab Cloud export file, if requested.
#   Read Google Invoice, if given, else use data from BillingDetails file.
#
###
if billing_details_file is not None and not args.skip_cloud:
    process_cloud_data()
    ilab_cloud_export_csv_dictwriter = open_ilab_output_dictwriter(output_subdir, "Cloud_Google")
else:
    ilab_cloud_export_csv_dictwriter = None

#####
#
# Output Consulting data into iLab Cluster export file, if requested.
#
####
if not args.skip_consulting:
    process_consulting_data()
    ilab_consulting_export_csv_dictwriter = open_ilab_output_dictwriter(output_subdir, "Consulting")
else:
    ilab_consulting_export_csv_dictwriter = None

# Write out cluster data to iLab export CSV file.
for pi_tag in sorted(pi_tag_list):

    print(" %s:" % pi_tag, end=' ')

    # Get the iLab service request ID for this PI.
    ilab_service_req = pi_tag_to_ilab_service_req_id[pi_tag]

    # If the PI explicitly is marked as not having a service request, skip them quietly.
    if str(ilab_service_req).lower() == 'n/a':
        print(" iLab service request not needed")
        continue

    # Get the cluster service level for this PI.
    service_level = pi_tag_to_service_level[pi_tag].lower()
    # Get the affiliation of this PI.
    affiliation = pi_tag_to_affiliation[pi_tag].lower()

    ###
    #
    # Write iLab export CSV file from output data structures.
    #
    ###

    #
    # Cluster Storage
    #
    if ilab_cluster_storage_export_csv_dictwriter is not None:

        # Output Cluster data into iLab Cluster export file, if requested.
        if service_level == 'free':
            print("free-tier", end=' ')

        elif service_level != 'no':  # service_level == 'full'
            # Storage

            # Get service IDs for base storage and additional storage
            (_, _, service_id_base_storage) = (
                get_rate_data_from_string(billing_config_wkbk, "Local HPC Storage", service_level,
                                          "Base Storage", affiliation))
            (_, _, service_id_addl_storage) = (
                get_rate_data_from_string(billing_config_wkbk, "Local HPC Storage", service_level,
                                          "Additional Storage", affiliation))

            if service_id_base_storage != "None":
                if output_ilab_csv_data_for_cluster_storage(ilab_cluster_storage_export_csv_dictwriter, pi_tag, ilab_service_req,
                                                            service_id_base_storage, service_id_addl_storage,
                                                            begin_month_timestamp, end_month_timestamp):
                    print("cluster-storage", end=' ')

    #
    # Cluster Compute
    #
    if ilab_cluster_compute_export_csv_dictwriter is not None:

        if service_level != "no":
            # Compute

            # Get service IDs for the lab's computing tier and the full computing tier (for non-lab members)
            (_, _, service_id_lab_computing) = (
                get_rate_data_from_string(billing_config_wkbk, "Local Computing", service_level,
                                          None, affiliation))
            (_, _, service_id_full_computing) = (
                get_rate_data_from_string(billing_config_wkbk, "Local Computing", "Full",
                                          None, affiliation))

            if output_ilab_csv_data_for_cluster_compute(ilab_cluster_compute_export_csv_dictwriter, pi_tag, ilab_service_req,
                                                        service_id_lab_computing,
                                                        service_id_full_computing,
                                                        begin_month_timestamp, end_month_timestamp):
                print("cluster-compute", end=' ')

    # Output Cloud data into iLab Cloud export file, if requested.
    if ilab_cloud_export_csv_dictwriter is not None:

        # Get service IDs for cloud services
        (_, _, service_id_cloud) = (
            get_rate_data_from_string(billing_config_wkbk, "Cloud Services", None,
                                      None, affiliation))

        if output_ilab_csv_data_for_cloud(ilab_cloud_export_csv_dictwriter, pi_tag, ilab_service_req,
                                          service_id_cloud,
                                          begin_month_timestamp, end_month_timestamp):
            print("cloud", end=' ')

    # Output Consulting data into iLab Cluster export file, if requested.
    if ilab_consulting_export_csv_dictwriter is not None:

        # Get service IDs for consulting
        (_, _, service_id_consulting_free) = (
            get_rate_data_from_string(billing_config_wkbk, "Bioinformatics Consulting", None,
                                      "Free Access", affiliation))
        (_, _, service_id_consulting_paid) = (
            get_rate_data_from_string(billing_config_wkbk, "Bioinformatics Consulting", None,
                                      None, affiliation))

        if output_ilab_csv_data_for_consulting(ilab_consulting_export_csv_dictwriter, pi_tag, ilab_service_req,
                                               service_id_consulting_free, service_id_consulting_paid,
                                               begin_month_timestamp, end_month_timestamp):
            print("consulting", end=' ')

    print()  # End line for PI tag.

print("iLab FILE CREATIONS COMPLETE!")
